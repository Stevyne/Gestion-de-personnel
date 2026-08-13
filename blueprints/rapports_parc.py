"""Exports PDF et Excel du parc matériel, limités au périmètre autorisé."""

import io

from flask import Blueprint, make_response
import openpyxl
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def creer_blueprint_rapports_parc(deps):
    bp = Blueprint('rapports_parc', __name__)
    db_cursor = deps['db_cursor']
    login_required = deps['login_required']
    department_scope_sql = deps['department_scope_sql']
    get_department_scope = deps['get_department_scope']

    def _donnees():
        with db_cursor() as (conn, cur):
            scope_sql, scope_params = department_scope_sql('d', 'nom', cur)
            cur.execute(f"""SELECT m.id,m.nom,m.categorie,d.nom AS departement,
                     m.quantite,m.seuil_alerte,m.unite,m.suivi_unitaire,
                     COALESCE(m.prix_acquisition,0) AS prix_acquisition,
                     COALESCE((SELECT SUM(a.quantite) FROM materiels_attributions a
                               WHERE a.materiel_id=m.id AND a.date_retour IS NULL),0) AS attribues,
                     (SELECT COUNT(*) FROM materiel_exemplaires ex
                       WHERE ex.materiel_id=m.id) AS exemplaires,
                     (SELECT COUNT(*) FROM materiel_exemplaires ex
                       WHERE ex.materiel_id=m.id AND ex.etat IN ('panne','reparation','rebut')) AS indisponibles
                    FROM materiels m LEFT JOIN departements d ON d.id=m.departement_id
                    WHERE {scope_sql} ORDER BY d.nom,m.nom""", scope_params)
            lignes = cur.fetchall()
            scope = get_department_scope(cur)
        return lignes, ('Tous les départements' if scope['is_global'] else
                        (scope.get('department') or 'Aucun département'))

    @bp.route('/export/materiels/pdf')
    @login_required
    def export_materiels_pdf():
        lignes, portee = _donnees()
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=24,
                                rightMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        elements = [Paragraph('Rapport du parc matériel', styles['Title']),
                    Paragraph(f'Périmètre : {portee}', styles['Normal']), Spacer(1, 12)]
        data = [['Article','Catégorie','Département','Stock','Attribué','Exemplaires','Indispo.','Valeur stock']]
        for row in lignes:
            data.append([row['nom'], row['categorie'] or '—', row['departement'] or '—',
                         f"{row['quantite']} {row['unite'] or ''}", row['attribues'],
                         row['exemplaires'], row['indisponibles'],
                         f"{float(row['prix_acquisition'] or 0) * row['quantite']:,.0f} Ar"])
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1e40af')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white), ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('GRID',(0,0),(-1,-1),.4,colors.HexColor('#cbd5e1')),
            ('FONTSIZE',(0,0),(-1,-1),8), ('VALIGN',(0,0),(-1,-1),'TOP'),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f8fafc')]),
        ]))
        elements.append(table)
        doc.build(elements)
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'attachment; filename=parc_materiel.pdf'
        return response

    @bp.route('/export/materiels/excel')
    @login_required
    def export_materiels_excel():
        lignes, portee = _donnees()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Parc matériel'
        sheet.append(['Périmètre', portee])
        sheet.append([])
        headers = ['Article','Catégorie','Département','Stock','Unité','Attribué',
                   'Exemplaires','Indisponibles','Prix unitaire','Valeur stock']
        sheet.append(headers)
        for cell in sheet[3]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1E40AF')
        for row in lignes:
            prix = float(row['prix_acquisition'] or 0)
            sheet.append([row['nom'],row['categorie'],row['departement'],row['quantite'],
                          row['unite'],row['attribues'],row['exemplaires'],row['indisponibles'],
                          prix,prix*row['quantite']])
        for column in sheet.columns:
            width = min(35, max(len(str(cell.value or '')) for cell in column) + 2)
            sheet.column_dimensions[column[0].column_letter].width = width
        buffer = io.BytesIO()
        workbook.save(buffer)
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=parc_materiel.xlsx'
        return response

    return bp

from flask import Flask, render_template, request, redirect, url_for, flash, session, make_response, jsonify, g
from werkzeug.utils import secure_filename
from psycopg2.extras import execute_values
import os
import secrets
import logging
from datetime import date, datetime, timedelta
from functools import wraps
import io
from urllib.parse import urlencode

from dotenv import load_dotenv
load_dotenv()   # charge .env avant la configuration des services

from flask_mail import Mail, Message

from extensions import db, migrate
from services.configuration import configurer_application
from services.database import creer_acces_postgres
from services.migrations import configurer_migrations
from services.security import configurer_securite_http
from services.common import (
    calculer_jours_acquis_prorata, calculer_retard, page_list, pagination_info,
)
from services.email_outbox import ajouter_email, traiter_outbox
from services.object_storage import object_storage
from services.observability import configure_logging, init_sentry, register_observability
from services.notifications import creer_service_notifications
from services.roles import GLOBAL_DATA_ROLES, ROLE_LABELS
from services.schema import initialiser_schema
from blueprints.absence_justifications import (
    ABSENCE_ACCEPTEE, ABSENCE_STATUT_LABELS, creer_blueprint_justifications,
)
from blueprints.absences import creer_blueprint_absences
from blueprints.messagerie import creer_blueprint_messagerie
from blueprints.notifications import creer_blueprint_notifications
from blueprints.recherche import (
    RECHERCHE_PAGES, creer_blueprint_recherche,  # noqa: F401 — réexport public
)
from blueprints.recrutement import creer_blueprint_recrutement
from blueprints.documents import creer_blueprint_documents
from blueprints.departements import creer_blueprint_departements
from blueprints.presences import creer_blueprint_presences
from blueprints.utilisateurs import creer_blueprint_utilisateurs
from blueprints.departs import creer_blueprint_departs
from blueprints.contrats import creer_blueprint_contrats
from blueprints.conges import creer_blueprint_conges
from blueprints.rapports_parc import creer_blueprint_rapports_parc
from blueprints.dashboards_roles import creer_blueprint_dashboards_roles
from blueprints.dashboard import creer_blueprint_dashboard
from blueprints.auth import creer_blueprint_auth
from blueprints.parc import (
    MAINTENANCE_OUVERTS,  # réexport de compatibilité pour les tests/plug-ins  # noqa: F401
    MAINTENANCE_VALIDATION_JOURS,
    creer_blueprint_parc,
)

# ==================== LOGGING / MONITORING ====================
configure_logging()
init_sentry()
logger = logging.getLogger('gestion_personnel')


# ==================== EXPORTS (PDF + EXCEL) ====================
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ==================== CONFIGURATION / DB ====================
DATABASE_URL = configurer_application(app, logger)
configurer_migrations(app, db, migrate, DATABASE_URL)
get_db, db_cursor, get_cursor = creer_acces_postgres(DATABASE_URL)

# ==================== UPLOADS (Documents) ====================
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'png', 'jpg', 'jpeg', 'txt'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Les photos de profil vivent dans un sous-dossier séparé des documents RH :
# elles sont servies publiquement par <img>, alors que les documents sont
# protégés par une route de téléchargement contrôlée.
AVATAR_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'avatars')
AVATAR_EXTENSIONS = {'png', 'jpg', 'jpeg'}
# 8 Mo à l'envoi : une photo prise au smartphone dépasse presque toujours 2 Mo.
# Elle est ensuite redimensionnée côté serveur, donc le fichier stocké reste petit.
AVATAR_MAX_BYTES = 8 * 1024 * 1024   # 8 Mo
AVATAR_MAX_SIDE = 512                # côté maximal de l'image stockée (px)
os.makedirs(AVATAR_FOLDER, exist_ok=True)

try:
    from PIL import Image, ImageOps
    PIL_DISPONIBLE = True
except ImportError:      # l'envoi reste possible, sans redimensionnement
    PIL_DISPONIBLE = False
    logger.warning("Pillow indisponible : les photos de profil ne seront pas redimensionnées.")

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Correspondance magic-bytes -> extensions cohérentes autorisées
_MAGIC_EXT = {
    b'%PDF': {'pdf'},
    b'\x89PNG\r\n\x1a\n': {'png'},
    b'\xff\xd8\xff': {'jpg', 'jpeg'},
    b'PK\x03\x04': {'docx', 'xlsx'},                 # Office Open XML (ZIP)
    b'PK\x05\x06': {'docx', 'xlsx'},                 # ZIP vide
    b'PK\x07\x08': {'docx', 'xlsx'},
    b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1': {'doc', 'xls'},  # OLE2 (anciens .doc/.xls)
}

def detect_file_type(fichier):
    """Détecte le vrai type du fichier via ses magic-bytes et renvoie
    l'extension (sans '.') cohérente avec le contenu parmi les types autorisés,
    ou None si le contenu ne correspond à rien de sûr."""
    stream = fichier.stream
    head = stream.read(8)
    stream.seek(0)
    candidates = None
    if head.startswith(b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'):
        candidates = {'doc', 'xls'}
    else:
        for magic, exts in _MAGIC_EXT.items():
            if head.startswith(magic):
                candidates = exts
                break
    if candidates is None:
        # Pas de magic connue : on accepte uniquement du texte UTF-8/ASCII
        try:
            chunk = stream.read(1024)
            stream.seek(0)
            chunk.decode('utf-8')
            candidates = {'txt'}
        except (UnicodeDecodeError, Exception):
            stream.seek(0)
            return None
    declared = fichier.filename.rsplit('.', 1)[-1].lower() if '.' in fichier.filename else ''
    return declared if declared in candidates else None


def get_admin_email():
    try:
        conn = get_db()
        cur = get_cursor(conn)
        cur.execute("SELECT email FROM employes WHERE LOWER(poste) LIKE '%admin%' OR LOWER(nom) LIKE '%admin%' OR email ILIKE '%admin%' LIMIT 1")
        row = cur.fetchone()
        cur.close()
        conn.close()
        if row and row.get('email'): return row['email']
    except: pass
    return app.config.get('ADMIN_EMAIL') or 'admin@entreprise.fr'

mail = Mail(app)


def queue_email(destinataire, sujet, corps_texte, corps_html=None,
                cur=None, event_key=None):
    """Place un e-mail dans l'outbox persistante.

    L'appel est un no-op assumé tant que ``EMAIL_ENABLED`` n'est pas activé.
    Si un curseur est fourni, le message appartient à la même transaction que
    l'événement métier ; sinon une courte transaction dédiée est ouverte.
    """
    if not app.config.get('EMAIL_ENABLED'):
        logger.info("[EMAIL DÉSACTIVÉ] → %s | %s", destinataire, sujet)
        return None
    if not destinataire:
        return None
    try:
        if cur is not None:
            # Une panne de l'outbox ne doit jamais annuler l'action métier.
            # Le SAVEPOINT restaure la transaction si l'INSERT e-mail échoue.
            cur.execute("SAVEPOINT ajout_email_outbox")
            try:
                resultat = ajouter_email(cur, destinataire, sujet, corps_texte,
                                          corps_html, event_key)
            except Exception:
                cur.execute("ROLLBACK TO SAVEPOINT ajout_email_outbox")
                cur.execute("RELEASE SAVEPOINT ajout_email_outbox")
                raise
            cur.execute("RELEASE SAVEPOINT ajout_email_outbox")
            return resultat
        with db_cursor(commit=True) as (conn, outbox_cur):
            return ajouter_email(outbox_cur, destinataire, sujet, corps_texte,
                                 corps_html, event_key)
    except Exception as exc:
        logger.error("Impossible d'ajouter l'e-mail à l'outbox: %s", exc,
                     exc_info=True)
        return None


def _send_outbox_message(message):
    """Adaptateur Flask-Mail appelé uniquement par le worker d'outbox."""
    msg = Message(
        subject=message['sujet'],
        recipients=[message['destinataire']],
        sender=app.config.get('MAIL_DEFAULT_SENDER'),
    )
    msg.body = message.get('corps_texte') or ''
    if message.get('corps_html'):
        msg.html = message['corps_html']
    mail.send(msg)


# ==================== SÉCURITÉ HTTP ====================
csrf, limiter, talisman = configurer_securite_http(app, logger)


# ==================== HTML EMAIL ====================
def send_html_email(recipients, subject, html_template, event_key=None, **context):
    """Rend un template puis met l'e-mail en file, sans SMTP dans la requête."""
    try:
        html_body = render_template(html_template, **context)
        liste = [recipients] if isinstance(recipients, str) else list(recipients or [])
        admin = get_admin_email()
        if admin and admin not in liste:
            liste.append(admin)  # équivalent de l'ancien CC administrateur
        for index, destinataire in enumerate(liste):
            queue_email(
                destinataire, subject,
                "Une notification de Gestion du Personnel vous attend. "
                "Consultez l'application pour les détails.",
                corps_html=html_body,
                event_key=f"{event_key}:{index}" if event_key else None,
            )
        return True
    except Exception as e:
        logger.error("Erreur préparation e-mail HTML: %s", e, exc_info=True)
        return False

HEURE_ARRIVEE_ATTENDUE = "09:00"

def get_role_label(role):
    return ROLE_LABELS.get(role, role or 'Employé')

# ==================== SESSIONS ACTIVES (présence + déconnexion à distance) ====================
# Durée d'inactivité au-delà de laquelle une session n'est plus considérée
# comme « en ligne » dans l'interface (elle reste valide tant qu'elle n'est
# ni expirée ni révoquée).
SESSION_ONLINE_WINDOW_MIN = int(os.environ.get('SESSION_ONLINE_WINDOW_MIN', 5))

def _client_ip():
    fwd = request.headers.get('X-Forwarded-For', '')
    if fwd:
        return fwd.split(',')[0].strip()[:45]
    return (request.remote_addr or '')[:45]


def enregistrer_session(user_id, username):
    """Inscrit la session courante au registre et renvoie son identifiant."""
    sid = secrets.token_urlsafe(32)
    try:
        with db_cursor(commit=True) as (conn, cur):
            cur.execute("""
                INSERT INTO sessions_actives (sid, user_id, username, ip_address, user_agent)
                VALUES (%s, %s, %s, %s, %s)
            """, (sid, user_id, username, _client_ip(),
                  (request.headers.get('User-Agent') or '')[:300]))
    except Exception as e:
        logger.error("Erreur enregistrer_session: %s", e, exc_info=True)
        return None
    return sid


def cloturer_session(sid, par=None):
    """Marque une session comme terminée (déconnexion volontaire ou révocation)."""
    if not sid:
        return
    try:
        with db_cursor(commit=True) as (conn, cur):
            cur.execute("""
                UPDATE sessions_actives
                   SET revoked_at = CURRENT_TIMESTAMP, revoked_by = %s
                 WHERE sid = %s AND revoked_at IS NULL
            """, (par, sid))
    except Exception as e:
        logger.error("Erreur cloturer_session: %s", e, exc_info=True)


def session_active():
    """Vrai si la session présentée par le navigateur est toujours valide.

    Rafraîchit `last_seen` au passage, ce qui alimente l'indicateur de présence.
    En cas d'incident base, on renvoie True : mieux vaut laisser passer que
    déconnecter tout le monde parce que la base est momentanément indisponible.
    """
    sid = session.get('sid')
    if not sid:
        # Session ouverte avant la mise en place du registre : on la tolère.
        return True
    try:
        with db_cursor(commit=True) as (conn, cur):
            cur.execute("SELECT revoked_at FROM sessions_actives WHERE sid = %s", (sid,))
            row = cur.fetchone()
            if row is None:
                return True          # inconnue (base réinitialisée) : on tolère
            if row['revoked_at'] is not None:
                return False         # révoquée par un administrateur
            cur.execute("UPDATE sessions_actives SET last_seen = CURRENT_TIMESTAMP WHERE sid = %s", (sid,))
    except Exception as e:
        logger.error("Erreur session_active: %s", e, exc_info=True)
    return True


def _refuser_session_revoquee():
    """Termine proprement une session dont l'accès vient d'être retiré."""
    session.clear()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        resp = make_response('', 204)
        resp.headers['X-Redirect-To'] = url_for('auth.login')
        return resp
    flash("Votre session a été fermée par un administrateur.", "warning")
    return redirect(url_for('auth.login'))


def role_required(*allowed_roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'user_id' not in session:
                flash('Veuillez vous connecter.', 'warning')
                return redirect(url_for('auth.login'))
            if not session_active():
                return _refuser_session_revoquee()
            role = session.get('role', 'employe')
            if role == 'admin' or role in allowed_roles:
                return f(*args, **kwargs)
            flash('Accès refusé.', 'danger')
            return redirect(url_for('dashboard.dashboard'))
        return decorated
    return decorator

# ==================== CLOISONNEMENT DES DONNÉES PAR DÉPARTEMENT =============
# Admin et RH ont une portée globale. Tous les autres rôles sont limités au
# département de leur fiche employé. Une absence de rattachement produit une
# portée vide — jamais un accès global implicite.


def get_department_scope(cur=None):
    """Portée courante, calculée depuis la base et mise en cache par requête.

    Le rôle n'est pas seulement lu dans le cookie de session : une promotion ou
    rétrogradation administrative prend ainsi effet dès la requête suivante.
    """
    if hasattr(g, 'department_scope'):
        return g.department_scope

    def charger(cursor):
        cursor.execute("""
            SELECT u.role, e.id AS employe_id, e.departement
              FROM users u LEFT JOIN employes e ON e.id = u.employe_id
             WHERE u.id = %s
        """, (session.get('user_id'),))
        return cursor.fetchone() or {}

    if cur is not None:
        row = charger(cur)
    else:
        with db_cursor() as (conn, cursor):
            row = charger(cursor)
    role = row.get('role') or 'employe'
    session['role'] = role
    session['role_label'] = get_role_label(role)
    if role in GLOBAL_DATA_ROLES:
        scope = {'is_global': True, 'department': None,
                 'employee_id': row.get('employe_id'), 'is_empty': False}
    else:
        departement = (row.get('departement') or '').strip() or None
        scope = {
            'is_global': False,
            'department': departement,
            'employee_id': row.get('employe_id'),
            'is_empty': not bool(departement),
        }
    g.department_scope = scope
    return scope


def department_scope_sql(alias='e', field='departement', cur=None):
    """Renvoie une condition SQL sûre et ses paramètres pour la portée courante."""
    scope = get_department_scope(cur)
    if scope['is_global']:
        return 'TRUE', []
    if scope['is_empty']:
        return 'FALSE', []
    return f"{alias}.{field} = %s", [scope['department']]


def employee_in_scope(cur, employe_id):
    """Vrai si la fiche employé appartient à la portée autorisée."""
    scope = get_department_scope(cur)
    if scope['is_global']:
        return True
    if not employe_id or scope['is_empty']:
        return False
    try:
        employe_id = int(employe_id)
    except (TypeError, ValueError):
        return False
    cur.execute("SELECT departement FROM employes WHERE id = %s", (employe_id,))
    row = cur.fetchone()
    return bool(row and row.get('departement') == scope['department'])


def department_id_in_scope(cur, departement_id):
    scope = get_department_scope(cur)
    if scope['is_global']:
        return True
    if not departement_id or scope['is_empty']:
        return False
    try:
        departement_id = int(departement_id)
    except (TypeError, ValueError):
        return False
    cur.execute("SELECT nom FROM departements WHERE id = %s", (departement_id,))
    row = cur.fetchone()
    return bool(row and row.get('nom') == scope['department'])


def _department_access_denied():
    logger.warning("Accès inter-département refusé: user=%s endpoint=%s",
                   session.get('username'), request.endpoint)
    if request.endpoint == 'recherche.api_recherche':
        return jsonify({'erreur': True, 'message': 'Accès refusé'}), 403
    flash("Accès refusé : cette donnée n'appartient pas à votre département.", "danger")
    return redirect(url_for('dashboard.dashboard'))


# Ressources identifiées par l'URL. Le SELECT renvoie toujours une colonne
# `departement`; le garde s'applique aux GET comme aux écritures.
DEPARTMENT_RESOURCE_GUARDS = {
    'view_employee': ("SELECT departement FROM employes WHERE id = %s", 'id'),
    'edit_employee': ("SELECT departement FROM employes WHERE id = %s", 'id'),
    'delete_employee': ("SELECT departement FROM employes WHERE id = %s", 'id'),
    'documents.add_employee_document': ("SELECT departement FROM employes WHERE id = %s", 'id'),
    'presences.clock_in': ("SELECT departement FROM employes WHERE id = %s", 'employe_id'),
    'presences.clock_out': ("SELECT departement FROM employes WHERE id = %s", 'employe_id'),
    'update_solde_conges': ("SELECT departement FROM employes WHERE id = %s", 'employe_id'),
    'presences.delete_presence': ("SELECT e.departement FROM presences p JOIN employes e ON e.id=p.employe_id WHERE p.id=%s", 'id'),
    'conges.avis_conge': ("SELECT e.departement FROM conges x JOIN employes e ON e.id=x.employe_id WHERE x.id=%s", 'id'),
    'conges.update_conge': ("SELECT e.departement FROM conges x JOIN employes e ON e.id=x.employe_id WHERE x.id=%s", 'id'),
    'conges.annuler_conge': ("SELECT e.departement FROM conges x JOIN employes e ON e.id=x.employe_id WHERE x.id=%s", 'id'),
    'conges.delete_conge': ("SELECT e.departement FROM conges x JOIN employes e ON e.id=x.employe_id WHERE x.id=%s", 'id'),
    'avis_permission': ("SELECT e.departement FROM permissions x JOIN employes e ON e.id=x.employe_id WHERE x.id=%s", 'id'),
    'update_permission': ("SELECT e.departement FROM permissions x JOIN employes e ON e.id=x.employe_id WHERE x.id=%s", 'id'),
    'annuler_permission': ("SELECT e.departement FROM permissions x JOIN employes e ON e.id=x.employe_id WHERE x.id=%s", 'id'),
    'delete_permission': ("SELECT e.departement FROM permissions x JOIN employes e ON e.id=x.employe_id WHERE x.id=%s", 'id'),
    'absences.delete_absence': ("SELECT e.departement FROM absences x JOIN employes e ON e.id=x.employe_id WHERE x.id=%s", 'id'),
    'documents.download_document': ("SELECT e.departement FROM documents x LEFT JOIN employes e ON e.id=x.employe_id WHERE x.id=%s", 'doc_id'),
    'documents.delete_document': ("SELECT e.departement FROM documents x LEFT JOIN employes e ON e.id=x.employe_id WHERE x.id=%s", 'doc_id'),
    'parc.edit_materiel': ("SELECT d.nom AS departement FROM materiels x LEFT JOIN departements d ON d.id=x.departement_id WHERE x.id=%s", 'id'),
    'parc.view_materiel': ("SELECT d.nom AS departement FROM materiels x LEFT JOIN departements d ON d.id=x.departement_id WHERE x.id=%s", 'id'),
    'parc.add_mouvement_materiel': ("SELECT d.nom AS departement FROM materiels x LEFT JOIN departements d ON d.id=x.departement_id WHERE x.id=%s", 'id'),
    'parc.attribuer_materiel': ("SELECT d.nom AS departement FROM materiels x LEFT JOIN departements d ON d.id=x.departement_id WHERE x.id=%s", 'id'),
    'parc.delete_materiel': ("SELECT d.nom AS departement FROM materiels x LEFT JOIN departements d ON d.id=x.departement_id WHERE x.id=%s", 'id'),
    'parc.add_exemplaire': ("SELECT d.nom AS departement FROM materiels x LEFT JOIN departements d ON d.id=x.departement_id WHERE x.id=%s", 'id'),
    'parc.etiquettes_materiel': ("SELECT d.nom AS departement FROM materiels x LEFT JOIN departements d ON d.id=x.departement_id WHERE x.id=%s", 'id'),
    'parc.retour_materiel': ("SELECT d.nom AS departement FROM materiels_attributions a JOIN materiels m ON m.id=a.materiel_id LEFT JOIN departements d ON d.id=m.departement_id WHERE a.id=%s", 'attribution_id'),
    'parc.accuser_attribution': ("SELECT d.nom AS departement FROM materiels_attributions a JOIN materiels m ON m.id=a.materiel_id LEFT JOIN departements d ON d.id=m.departement_id WHERE a.id=%s", 'attribution_id'),
    'parc.relancer_accuse': ("SELECT d.nom AS departement FROM materiels_attributions a JOIN materiels m ON m.id=a.materiel_id LEFT JOIN departements d ON d.id=m.departement_id WHERE a.id=%s", 'attribution_id'),
    'parc.view_inventaire': ("SELECT d.nom AS departement FROM inventaires x LEFT JOIN departements d ON d.id=x.departement_id WHERE x.id=%s", 'id'),
    'parc.compter_inventaire': ("SELECT d.nom AS departement FROM inventaires x LEFT JOIN departements d ON d.id=x.departement_id WHERE x.id=%s", 'id'),
    'parc.cloturer_inventaire': ("SELECT d.nom AS departement FROM inventaires x LEFT JOIN departements d ON d.id=x.departement_id WHERE x.id=%s", 'id'),
    'parc.annuler_inventaire': ("SELECT d.nom AS departement FROM inventaires x LEFT JOIN departements d ON d.id=x.departement_id WHERE x.id=%s", 'id'),
    'parc.view_exemplaire': ("SELECT d.nom AS departement FROM materiel_exemplaires ex JOIN materiels m ON m.id=ex.materiel_id LEFT JOIN departements d ON d.id=m.departement_id WHERE ex.id=%s", 'id'),
    'parc.edit_exemplaire': ("SELECT d.nom AS departement FROM materiel_exemplaires ex JOIN materiels m ON m.id=ex.materiel_id LEFT JOIN departements d ON d.id=m.departement_id WHERE ex.id=%s", 'id'),
    'parc.delete_exemplaire': ("SELECT d.nom AS departement FROM materiel_exemplaires ex JOIN materiels m ON m.id=ex.materiel_id LEFT JOIN departements d ON d.id=m.departement_id WHERE ex.id=%s", 'id'),
    'parc.signaler_panne': ("SELECT d.nom AS departement FROM materiel_exemplaires ex JOIN materiels m ON m.id=ex.materiel_id LEFT JOIN departements d ON d.id=m.departement_id WHERE ex.id=%s", 'id'),
    'parc.discussion_maintenance': ("SELECT d.nom AS departement FROM materiel_maintenances mt JOIN materiel_exemplaires ex ON ex.id=mt.exemplaire_id JOIN materiels m ON m.id=ex.materiel_id LEFT JOIN departements d ON d.id=m.departement_id WHERE mt.id=%s", 'id'),
    'parc.assigner_maintenance': ("SELECT d.nom AS departement FROM materiel_maintenances mt JOIN materiel_exemplaires ex ON ex.id=mt.exemplaire_id JOIN materiels m ON m.id=ex.materiel_id LEFT JOIN departements d ON d.id=m.departement_id WHERE mt.id=%s", 'id'),
    'parc.envoyer_maintenance': ("SELECT d.nom AS departement FROM materiel_maintenances mt JOIN materiel_exemplaires ex ON ex.id=mt.exemplaire_id JOIN materiels m ON m.id=ex.materiel_id LEFT JOIN departements d ON d.id=m.departement_id WHERE mt.id=%s", 'id'),
    'parc.cloturer_maintenance': ("SELECT d.nom AS departement FROM materiel_maintenances mt JOIN materiel_exemplaires ex ON ex.id=mt.exemplaire_id JOIN materiels m ON m.id=ex.materiel_id LEFT JOIN departements d ON d.id=m.departement_id WHERE mt.id=%s", 'id'),
    'parc.valider_maintenance': ("SELECT d.nom AS departement FROM materiel_maintenances mt JOIN materiel_exemplaires ex ON ex.id=mt.exemplaire_id JOIN materiels m ON m.id=ex.materiel_id LEFT JOIN departements d ON d.id=m.departement_id WHERE mt.id=%s", 'id'),
    'parc.annuler_maintenance': ("SELECT d.nom AS departement FROM materiel_maintenances mt JOIN materiel_exemplaires ex ON ex.id=mt.exemplaire_id JOIN materiels m ON m.id=ex.materiel_id LEFT JOIN departements d ON d.id=m.departement_id WHERE mt.id=%s", 'id'),
    'parc.etiquettes_departement': ("SELECT nom AS departement FROM departements WHERE id=%s", 'id'),
    'parc.materiels_departement': ("SELECT nom AS departement FROM departements WHERE id=%s", 'id'),
    'auth.avatar_image': ("SELECT e.departement FROM users u LEFT JOIN employes e ON e.id=u.employe_id WHERE u.photo=%s", 'filename'),
    # Le Blueprint des justificatifs applique une règle plus stricte que le
    # département (propriétaire ou RH) et conserve ses réponses 404/403.
}

FORM_EMPLOYEE_SCOPE_FIELDS = {
    'presences.presences': 'quick_employe_id', 'presences.add_presence': 'employe_id',
    'conges.add_conge': 'employe_id', 'add_permission': 'employe_id',
    'absences.add_absence': 'employe_id', 'documents.documents': 'employe_id',
    'parc.attribuer_materiel': 'employe_id', 'parc.add_mouvement_materiel': 'employe_id',
    'parc.edit_exemplaire': 'employe_id',
}
FORM_DEPARTMENT_SCOPE_FIELDS = {
    'parc.add_materiel': 'departement_id', 'parc.edit_materiel': 'departement_id',
    'parc.add_inventaire': 'departement_id',
}
FORM_USER_SCOPE_FIELDS = {'parc.assigner_maintenance': 'assigne_user_id'}


@app.before_request
def enforce_department_resource_scope():
    """Garde central anti-contournement pour URL et formulaires forgés."""
    if not session.get('user_id'):
        return None
    scope = get_department_scope()
    if scope['is_global']:
        return None

    # Les prestataires sont un référentiel global sans département : seuls
    # admin/RH peuvent le consulter ou le modifier.
    if request.endpoint in ('parc.prestataires_page', 'parc.basculer_prestataire'):
        return _department_access_denied()

    guard = DEPARTMENT_RESOURCE_GUARDS.get(request.endpoint)
    if guard and request.view_args:
        query, key = guard
        value = request.view_args.get(key)
        if value is not None:
            with db_cursor() as (conn, cur):
                cur.execute(query, (value,))
                row = cur.fetchone()
            if row and (scope['is_empty'] or row.get('departement') != scope['department']):
                return _department_access_denied()

    if request.method == 'POST':
        employee_field = FORM_EMPLOYEE_SCOPE_FIELDS.get(request.endpoint)
        raw_employee = request.form.get(employee_field) if employee_field else None
        if raw_employee:
            with db_cursor() as (conn, cur):
                if not employee_in_scope(cur, raw_employee):
                    return _department_access_denied()
        department_field = FORM_DEPARTMENT_SCOPE_FIELDS.get(request.endpoint)
        raw_department = request.form.get(department_field) if department_field else None
        if raw_department:
            with db_cursor() as (conn, cur):
                if not department_id_in_scope(cur, raw_department):
                    return _department_access_denied()
        user_field = FORM_USER_SCOPE_FIELDS.get(request.endpoint)
        raw_user = request.form.get(user_field) if user_field else None
        if raw_user:
            try:
                raw_user = int(raw_user)
            except (TypeError, ValueError):
                return _department_access_denied()
            with db_cursor() as (conn, cur):
                user_scope, user_params = department_scope_sql('e', 'departement', cur)
                cur.execute(f"""SELECT 1 FROM users u JOIN employes e ON e.id=u.employe_id
                                WHERE u.id=%s AND {user_scope}""",
                            [raw_user] + user_params)
                if not cur.fetchone():
                    return _department_access_denied()
    return None


# ==================== NOTIFICATIONS (Base de données - support multi-utilisateur réel) ====================
# ==================== NOTIFICATIONS ====================
(create_notification, get_unread_notifications, mark_all_read,
 get_all_notifications) = creer_service_notifications(db_cursor, logger)


@app.after_request
def modal_redirect_passthrough(response):
    """Ne pas suivre les redirections pour les soumissions de la popup.

    Le JS de la fenêtre modale poste en AJAX (`fetch`) avec l'en-tête
    X-Requested-With. Si le navigateur suit lui-même la redirection 302,
    la page de destination est rendue dans le fetch : les messages flash
    sont alors consommés et « perdus » avant que le navigateur ne recharge
    réellement la liste.

    On transforme donc la redirection en réponse 204 portant l'URL cible
    dans l'en-tête X-Redirect-To ; le JS lit cet en-tête et navigue
    lui-même, ce qui laisse le flash intact pour le vrai chargement.

    Les formulaires du panneau d'activité sont explicitement exclus : leur JS
    a besoin de suivre la redirection vers un nouveau fragment HTML.
    """
    try:
        if (response.status_code in (301, 302, 303, 307, 308)
                and request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                # Le panneau latéral suit lui-même la redirection afin de
                # remplacer son fragment. Le 204 est réservé aux popups CRUD.
                and request.headers.get('X-Activity-Panel') != '1'
                and request.method == 'POST'):
            target = response.headers.get('Location')
            if target:
                resp = make_response('', 204)
                resp.headers['X-Redirect-To'] = target
                return resp
    except Exception:
        pass
    return response


@app.context_processor
def inject_context():
    try:
        user_id = session.get('user_id')
        unread_count = len(get_unread_notifications(user_id)) if user_id else 0
    except:
        unread_count = 0
    # Mode « popup » : quand une page de formulaire est demandée avec ?modal=1
    # (par le JS d'ouverture de la fenêtre modale), on l'affiche via un layout
    # réduit au seul contenu, sans <html>/navigation/pied de page.
    # Les templates de formulaire font `{% extends layout %}`, ce qui évite de
    # dupliquer les formulaires entre la page classique et la popup.
    is_modal = request.args.get('modal') == '1'
    # Photo de profil du compte connecté, pour l'avatar de la barre de navigation.
    photo_profil = None
    try:
        if session.get('user_id'):
            with db_cursor() as (conn, cur):
                cur.execute("SELECT photo FROM users WHERE id = %s", (session['user_id'],))
                row = cur.fetchone()
                photo_profil = row['photo'] if row else None
    except Exception:
        photo_profil = None
    return {
        'unread_notifications': unread_count,
        'current_role': session.get('role', 'employe'),
        'role_label': session.get('role_label') or get_role_label(session.get('role', 'employe')),
        'is_modal': is_modal,
        'layout': '_modal_layout.html' if is_modal else 'base.html',
        'session_online_window': SESSION_ONLINE_WINDOW_MIN,
        'photo_profil': photo_profil,
    }

# ==================== RETARD EMAIL (HTML) ====================
def send_retard_email(employee_name, employee_email, retard_minutes, date_str, heure_arrivee):
    admin_email = get_admin_email()
    if not app.config.get('EMAIL_ENABLED'):
        logger.info(f"[EMAIL DÉSACTIVÉ] → {employee_name} +{retard_minutes} min")
        return True
    try:
        subject = f"⚠️ Retard détecté - {employee_name}"
        return send_html_email(
            recipients=[employee_email] if employee_email else [admin_email],
            subject=subject,
            html_template="emails/retard.html",
            event_key=f"retard:{employee_email or employee_name}:{date_str}:{heure_arrivee}",
            prenom=employee_name.split()[0] if employee_name else "Employé",
            nom_complet=employee_name,
            date_str=date_str,
            heure_arrivee=heure_arrivee,
            retard_minutes=retard_minutes,
            heure_attendue=HEURE_ARRIVEE_ATTENDUE,
            admin_name="Administrateur Système"
        )
    except Exception as e:
        logger.error("Erreur mise en file e-mail retard: %s", e, exc_info=True)
        return False

# Services de pagination extraits dans services/common.py.
def log_action(user_id=None, username=None, action="", entity_type=None, entity_id=None, details=None):
    try:
        conn = get_db()
        cur = get_cursor(conn)
        ip = getattr(request, 'remote_addr', None)
        cur.execute('INSERT INTO audit_logs (user_id, username, action, entity_type, entity_id, details, ip_address) VALUES (%s,%s,%s,%s,%s,%s,%s)',
                    (user_id, username, action, entity_type, entity_id, details, ip))
        conn.commit()
        cur.close(); conn.close()
    except: pass

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Veuillez vous connecter.', 'warning')
            return redirect(url_for('auth.login'))
        # Une session révoquée à distance est rejetée dès la requête suivante.
        if not session_active():
            return _refuser_session_revoquee()
        return f(*args, **kwargs)
    return decorated

# calculer_retard est réexporté depuis services/common.py.
def get_current_user_row():
    """Renvoie la ligne `users` du compte connecté (identifiant, rôle, photo…)."""
    if 'user_id' not in session:
        return None
    try:
        with db_cursor() as (conn, cur):
            cur.execute("SELECT id, username, role, employe_id, photo FROM users WHERE id = %s",
                        (session['user_id'],))
            return cur.fetchone()
    except Exception as e:
        logger.error("Erreur get_current_user_row: %s", e, exc_info=True)
        return None


def enregistrer_photo_profil(fichier, user_id):
    """Valide puis enregistre une photo de profil.

    Renvoie (nom_du_fichier, None, contenu_bytes) en cas de succès, sinon
    (None, message, None). La validation porte sur le CONTENU réel
    (magic-bytes) et pas seulement sur l'extension : un script renommé en
    .png est rejeté.

    Le fichier est toujours écrit sur disque (cache pour la durée de vie de
    l'instance) ET ses octets sont renvoyés pour être stockés en base
    (persistant) par l'appelant — voir la remarque sur le disque éphémère
    dans `job_alertes_expiration_documents` / `_traiter_upload_document`.
    """
    if not fichier or not fichier.filename:
        return None, "Aucune image sélectionnée.", None

    ext = fichier.filename.rsplit('.', 1)[-1].lower() if '.' in fichier.filename else ''
    if ext not in AVATAR_EXTENSIONS:
        return None, "Format non accepté. Utilisez une image PNG ou JPEG.", None

    # Taille : on mesure sans charger tout le fichier en mémoire.
    fichier.stream.seek(0, os.SEEK_END)
    taille = fichier.stream.tell()
    fichier.stream.seek(0)
    if taille > AVATAR_MAX_BYTES:
        return None, (f"Image trop volumineuse ({taille // (1024 * 1024)} Mo). "
                      f"Maximum : {AVATAR_MAX_BYTES // (1024 * 1024)} Mo."), None
    if taille == 0:
        return None, "Le fichier est vide.", None

    if detect_file_type(fichier) not in AVATAR_EXTENSIONS:
        return None, "Ce fichier n'est pas une image valide.", None

    # Nom imprévisible : empêche de deviner l'URL de la photo d'autrui.
    nom = f"u{user_id}_{secrets.token_hex(8)}.{ 'jpg' if ext == 'jpeg' else ext }"
    chemin = os.path.join(AVATAR_FOLDER, secure_filename(nom))

    if not PIL_DISPONIBLE:
        fichier.stream.seek(0)
        contenu = fichier.stream.read()
        with open(chemin, 'wb') as f:
            f.write(contenu)
        return nom, None, contenu

    # Redimensionnement : l'image est recadrée en carré puis réduite. Le
    # ré-encodage par Pillow supprime au passage toute charge utile cachée
    # dans le fichier d'origine (métadonnées EXIF, données après l'image).
    try:
        fichier.stream.seek(0)
        with Image.open(fichier.stream) as img:
            img = ImageOps.exif_transpose(img)          # respecte l'orientation photo
            img = ImageOps.fit(img, (AVATAR_MAX_SIDE, AVATAR_MAX_SIDE),
                               method=Image.LANCZOS, centering=(0.5, 0.4))
            buffer = io.BytesIO()
            if nom.endswith('.png'):
                img.convert('RGBA').save(buffer, 'PNG', optimize=True)
            else:
                img.convert('RGB').save(buffer, 'JPEG', quality=88, optimize=True)
            contenu = buffer.getvalue()
            with open(chemin, 'wb') as f:
                f.write(contenu)
    except Exception as e:
        logger.error("Erreur de traitement de la photo de profil : %s", e, exc_info=True)
        return None, "Cette image n'a pas pu être traitée. Essayez un autre fichier.", None

    return nom, None, contenu


def supprimer_photo_profil(nom_fichier):
    """Efface le fichier d'une ancienne photo, sans jamais interrompre l'appelant."""
    if not nom_fichier:
        return
    try:
        chemin = os.path.join(AVATAR_FOLDER, secure_filename(nom_fichier))
        # Garde-fou : on ne supprime rien en dehors du dossier des avatars.
        if os.path.dirname(os.path.abspath(chemin)) == os.path.abspath(AVATAR_FOLDER) \
                and os.path.isfile(chemin):
            os.remove(chemin)
    except Exception as e:
        logger.error("Erreur supprimer_photo_profil: %s", e, exc_info=True)


def get_current_employee():
    if 'user_id' not in session: return None
    try:
        with db_cursor() as (conn, cur):
            cur.execute("SELECT e.* FROM employes e JOIN users u ON u.employe_id = e.id WHERE u.id = %s LIMIT 1", (session['user_id'],))
            return cur.fetchone()
    except:
        return None

Q_SOLDE = ("SELECT jours_acquis, jours_utilises FROM soldes_conges "
           "WHERE employe_id = %s AND annee = %s")

# ==================== WORKFLOWS RH (congés / permissions) ====================
# Circuit : l'employé dépose → le manager de son département donne un avis →
# le RH tranche. Les statuts historiques ('en attente', 'approuvé', 'refusé')
# sont conservés tels quels pour ne pas invalider les demandes déjà en base.
DEMANDE_EN_ATTENTE = 'en attente'      # déposée, avis manager non rendu
DEMANDE_AVIS_RENDU = 'avis rendu'      # le manager s'est prononcé, RH à décider
DEMANDE_APPROUVEE  = 'approuvé'
DEMANDE_REFUSEE    = 'refusé'
DEMANDE_ANNULEE    = 'annulé'
DEMANDE_OUVERTES   = (DEMANDE_EN_ATTENTE, DEMANDE_AVIS_RENDU)

DEMANDE_LIBELLES = {
    DEMANDE_EN_ATTENTE: "En attente d'avis du manager",
    DEMANDE_AVIS_RENDU: "Avis rendu — décision RH attendue",
    DEMANDE_APPROUVEE:  "Approuvé",
    DEMANDE_REFUSEE:    "Refusé",
    DEMANDE_ANNULEE:    "Annulé",
}


def _peut_decider_rh():
    """Décision finale sur une demande de congé / permission."""
    return session.get('role') in ('admin', 'rh')


def _peut_donner_avis():
    """Étape intermédiaire : réservée aux managers (l'admin dépanne)."""
    return session.get('role') in ('admin', 'manager')


def _managers_du_departement(cur, departement):
    """Comptes managers rattachés à un département (lien par nom).

    Le champ `departements.responsable` est un texte libre, inexploitable pour
    router une demande : on passe donc par les comptes de rôle `manager` dont
    l'employé lié appartient au même département.
    """
    if not departement:
        return []
    cur.execute("""
        SELECT u.id, u.username, e.email FROM users u
          JOIN employes e ON e.id = u.employe_id
         WHERE u.role = 'manager' AND e.departement = %s
    """, (departement,))
    return cur.fetchall()


def _notifier_roles(cur, roles, titre, message, type_='info', sauf=None):
    """Notifie tous les comptes portant l'un des rôles donnés."""
    cur.execute("SELECT id FROM users WHERE role IN %s", (tuple(roles),))
    for row in cur.fetchall():
        if sauf and row['id'] == sauf:
            continue
        create_notification(row['id'], titre, message, type_, cur=cur)


def _envoyer_roles(cur, roles, sujet, message, cle_prefixe):
    """Met en file un e-mail pour chaque compte de rôle disposant d'une adresse."""
    cur.execute("""
        SELECT u.id, e.email FROM users u
        LEFT JOIN employes e ON e.id = u.employe_id
        WHERE u.role IN %s AND e.email IS NOT NULL
    """, (tuple(roles),))
    for row in cur.fetchall():
        queue_email(row['email'], sujet, message, cur=cur,
                    event_key=f"{cle_prefixe}:{row['id']}")


def _notifier_employe_evenement(cur, employe_id, titre, message,
                                 type_='info', cle_evenement=None):
    """Notification interne + e-mail éventuel à l'employé concerné."""
    cur.execute("""
        SELECT e.email, u.id AS user_id
          FROM employes e LEFT JOIN users u ON u.employe_id = e.id
         WHERE e.id = %s ORDER BY u.id NULLS LAST LIMIT 1
    """, (employe_id,))
    cible = cur.fetchone()
    if not cible:
        return
    if cible.get('user_id'):
        create_notification(cible['user_id'], titre, message, type_, cur=cur)
    queue_email(cible.get('email'), titre, message, cur=cur,
                event_key=cle_evenement)


def _user_id_de_employe(cur, employe_id):
    """Compte utilisateur lié à un employé, s'il en a un."""
    if not employe_id:
        return None
    cur.execute("SELECT id FROM users WHERE employe_id = %s LIMIT 1", (employe_id,))
    row = cur.fetchone()
    return row['id'] if row else None


def _libelle_employe(cur, employe_id):
    cur.execute("SELECT nom, prenom, departement FROM employes WHERE id = %s", (employe_id,))
    e = cur.fetchone()
    if not e:
        return "?", None
    return f"{e['prenom']} {e['nom']}".strip(), e['departement']


# ==================== SOLDES DE CONGÉS (Phase 2) ====================

# Calcul proratisé réexporté depuis services/common.py.

def get_solde_conges(employe_id, annee=None):
    """Retourne le solde de congés d'un employé (jours acquis, utilisés, restants)"""
    if annee is None:
        annee = datetime.now().year
    try:
        with db_cursor(commit=True) as (conn, cur):
            cur.execute("""
                SELECT * FROM soldes_conges 
                WHERE employe_id = %s AND annee = %s
            """, (employe_id, annee))
            solde = cur.fetchone()

            if not solde:
                cur.execute("SELECT date_embauche FROM employes WHERE id = %s", (employe_id,))
                emp_row = cur.fetchone()
                acquis_initial = calculer_jours_acquis_prorata(
                    emp_row.get('date_embauche') if emp_row else None, annee
                )
                cur.execute("""
                    INSERT INTO soldes_conges (employe_id, annee, jours_acquis, jours_utilises)
                    VALUES (%s, %s, %s, 0)
                    RETURNING *
                """, (employe_id, annee, acquis_initial))
                solde = cur.fetchone()

            acquis = float(solde.get('jours_acquis') or 25)
            utilises = float(solde.get('jours_utilises') or 0)
            return {
                'jours_acquis': acquis,
                'jours_utilises': utilises,
                'jours_restants': round(acquis - utilises, 1),
                'jours_acquis_manuel': bool(solde.get('jours_acquis_manuel')),
                'annee': annee
            }
    except Exception as e:
        logger.error("Erreur get_solde_conges: %s", e, exc_info=True)
        return {'jours_acquis': 25, 'jours_utilises': 0, 'jours_restants': 25,
                'jours_acquis_manuel': False, 'annee': annee}


def mettre_a_jour_solde(employe_id, jours_delta, annee=None):
    """Ajoute ou soustrait des jours du solde (appelé lors de l'approbation/refus)"""
    if annee is None:
        annee = datetime.now().year
    try:
        conn = get_db()
        cur = get_cursor(conn)
        cur.execute("SELECT date_embauche FROM employes WHERE id = %s", (employe_id,))
        emp_row = cur.fetchone()
        acquis_initial = calculer_jours_acquis_prorata(
            emp_row.get('date_embauche') if emp_row else None, annee
        )
        cur.execute("""
            INSERT INTO soldes_conges (employe_id, annee, jours_acquis, jours_utilises)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (employe_id, annee) 
            DO UPDATE SET jours_utilises = GREATEST(0, soldes_conges.jours_utilises + %s)
        """, (employe_id, annee, acquis_initial, jours_delta, jours_delta))
        conn.commit()
        cur.close()
        conn.close()
        return True
    except Exception as e:
        logger.error("Erreur mise à jour solde: %s", e, exc_info=True)
        return False


def recalculer_solde(employe_id, annee=None, cur=None):
    """
    Recalcule automatiquement les jours utilisés depuis les congés approuvés.

    Si `cur` est fourni, réutilise ce curseur/cette transaction (nécessaire quand
    on est appelé depuis update_conge : la connexion séparée qu'on ouvrait ici
    avant ne voyait pas l'UPDATE du statut pas encore commité par l'appelant,
    en isolation READ COMMITTED — le solde ne se mettait donc jamais à jour).
    """
    if annee is None:
        annee = datetime.now().year

    def _do(cur):
        cur.execute("""
            SELECT COALESCE(SUM(nombre_jours), 0) as total
            FROM conges 
            WHERE employe_id = %s 
              AND statut = 'approuvé'
              AND type_conge = 'congé payé'
              AND EXTRACT(YEAR FROM date_debut) = %s
        """, (employe_id, annee))
        total = float(cur.fetchone()['total'] or 0)

        cur.execute("SELECT date_embauche FROM employes WHERE id = %s", (employe_id,))
        emp_row = cur.fetchone()
        acquis_initial = calculer_jours_acquis_prorata(
            emp_row.get('date_embauche') if emp_row else None, annee
        )
        cur.execute("""
            INSERT INTO soldes_conges (employe_id, annee, jours_acquis, jours_utilises)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (employe_id, annee) 
            DO UPDATE SET jours_utilises = %s
        """, (employe_id, annee, acquis_initial, total, total))
        return total

    try:
        if cur is not None:
            return _do(cur)
        with db_cursor(commit=True) as (conn, cur):
            return _do(cur)
    except Exception as e:
        logger.error("Erreur recalcul solde: %s", e, exc_info=True)
        return 0


def init_db():
    """Réexport de compatibilité du bootstrap déplacé dans services/schema.py."""
    return initialiser_schema(
        get_db, get_cursor, logger, calculer_jours_acquis_prorata
    )
# ==================== AUTH ====================
# ==================== AUTHENTIFICATION ======================================
# Routes extraites dans blueprints/auth.py.

# Recherche globale extraite dans blueprints/recherche.py.

# ==================== ESPACE PERSONNEL =====================================
# Routes extraites dans blueprints/auth.py.

# ==================== SELF-SERVICE ====================
@app.route('/self-service')
@app.route('/mon-espace')
@login_required
def self_service():
    emp = get_current_employee()
    my_presences = []
    my_conges = []
    my_absences = []
    mon_solde = None
    materiels_a_confirmer = 0
    if emp:
        conn = get_db()
        cur = get_cursor(conn)
        cur.execute("SELECT * FROM presences WHERE employe_id = %s ORDER BY date DESC LIMIT 30", (emp['id'],))
        my_presences = cur.fetchall()
        for p in my_presences: p['retard_minutes'] = calculer_retard(p['heure_arrivee'])
        cur.execute("SELECT * FROM conges WHERE employe_id = %s ORDER BY date_demande DESC LIMIT 15", (emp['id'],))
        my_conges = cur.fetchall()
        cur.execute("""SELECT id, date, statut, motif_refus FROM absences
                       WHERE employe_id = %s ORDER BY date DESC LIMIT 8""",
                    (emp['id'],))
        my_absences = cur.fetchall()
        mon_solde = get_solde_conges(emp['id'])
        # Nombre d'équipements dont l'employé n'a pas encore accusé réception :
        # sert à afficher un rappel visible dès l'accueil de l'espace.
        cur.execute("""SELECT COUNT(*) AS nb FROM materiels_attributions
                        WHERE employe_id = %s AND accuse_reception = FALSE""", (emp['id'],))
        materiels_a_confirmer = cur.fetchone()['nb']
        cur.close(); conn.close()
    return render_template('self_service.html', employee=emp, my_presences=my_presences,
                           my_conges=my_conges, my_absences=my_absences,
                           absence_statut_labels=ABSENCE_STATUT_LABELS,
                           mon_solde=mon_solde, libelles=DEMANDE_LIBELLES,
                           materiels_a_confirmer=materiels_a_confirmer)

@app.route('/self-service/presences')
@login_required
def self_service_presences():
    emp = get_current_employee()
    if not emp:
        flash("Aucun employé lié à votre compte.", "warning")
        return redirect(url_for('self_service'))
    conn = get_db()
    cur = get_cursor(conn)
    cur.execute("SELECT * FROM presences WHERE employe_id = %s ORDER BY date DESC", (emp['id'],))
    presences = cur.fetchall()
    for p in presences: p['retard_minutes'] = calculer_retard(p['heure_arrivee'])
    cur.close(); conn.close()
    return render_template('self_service_presences.html', presences=presences, employee=emp)

@app.route('/self-service/conges')
@login_required
def self_service_conges():
    """Mes demandes de congés et de permissions, avec dépôt et annulation."""
    emp = get_current_employee()
    if not emp:
        flash("Aucun employé lié à votre compte.", "warning")
        return redirect(url_for('self_service'))
    with db_cursor() as (conn, cur):
        cur.execute("SELECT * FROM conges WHERE employe_id = %s "
                    "ORDER BY date_demande DESC", (emp['id'],))
        conges = cur.fetchall()
        cur.execute("SELECT * FROM permissions WHERE employe_id = %s "
                    "ORDER BY date_debut DESC", (emp['id'],))
        permissions_list = cur.fetchall()
    return render_template('self_service_conges.html', conges=conges,
                           permissions=permissions_list, employee=emp,
                           mon_solde=get_solde_conges(emp['id']),
                           libelles=DEMANDE_LIBELLES,
                           ouvertes=DEMANDE_OUVERTES)


@app.route('/self-service/materiels')
@login_required
def self_service_materiels():
    """Matériel qui m'est attribué, et accusés de réception en attente."""
    emp = get_current_employee()
    if not emp:
        flash("Aucun employé lié à votre compte.", "warning")
        return redirect(url_for('self_service'))
    with db_cursor() as (conn, cur):
        cur.execute("""
            SELECT a.*, m.nom AS materiel_nom, m.categorie, m.unite,
                   ex.numero_inventaire AS exemplaire_numero
              FROM materiels_attributions a
              JOIN materiels m ON m.id = a.materiel_id
              LEFT JOIN materiel_exemplaires ex ON ex.id = a.exemplaire_id
             WHERE a.employe_id = %s
             ORDER BY a.accuse_reception ASC, a.date_attribution DESC
        """, (emp['id'],))
        attributions = cur.fetchall()
    return render_template('self_service_materiels.html',
                           attributions=attributions, employee=emp)


# ==================== EXPORTS ====================
def create_presences_pdf(data, title="Rapport des Présences"):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=18, textColor=colors.HexColor('#1e40af'))
    elements = [Paragraph(title, title_style), Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']), Spacer(1, 12)]
    if data:
        tdata = [["Date", "Employé", "Arrivée", "Retard", "Départ", "Statut"]]
        for row in data:
            ret = calculer_retard(row.get('heure_arrivee'))
            nom = f"{row.get('prenom','')} {row.get('nom','')}".strip()
            tdata.append([str(row.get('date','')), nom, str(row.get('heure_arrivee') or '—')[:5], f"+{ret} min" if ret > 0 else "—", str(row.get('heure_depart') or '—')[:5], row.get('statut','')])
        t = Table(tdata, colWidths=[2.3*cm,5*cm,2*cm,2.1*cm,2*cm,2.5*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    return buffer

def create_conges_pdf(data, title="Rapport des Congés"):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=18, textColor=colors.HexColor('#166534'))
    elements = [Paragraph(title, title_style), Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']), Spacer(1, 12)]
    if data:
        tdata = [["Employé", "Type", "Début", "Fin", "Jours", "Statut"]]
        for row in data:
            nom = f"{row.get('prenom','')} {row.get('nom','')}".strip()
            tdata.append([nom, row.get('type_conge',''), str(row.get('date_debut','')), str(row.get('date_fin','')), str(row.get('nombre_jours','')), row.get('statut','')])
        t = Table(tdata, colWidths=[5*cm,3.3*cm,2.7*cm,2.7*cm,1.4*cm,2.4*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#166534')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    return buffer

def create_presences_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Présences"
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    headers = ["Date", "Employé", "Arrivée", "Retard (min)", "Départ", "Statut"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    for i, row in enumerate(data, 2):
        ret = calculer_retard(row.get('heure_arrivee'))
        nom = f"{row.get('prenom','')} {row.get('nom','')}".strip()
        vals = [str(row.get('date','')), nom, str(row.get('heure_arrivee') or '')[:5], ret if ret > 0 else 0, str(row.get('heure_depart') or '')[:5], row.get('statut','')]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = thin
            if c == 4 and v > 0: cell.font = Font(color="DC2626", bold=True)
    for c in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def create_conges_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Congés"
    header_fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    headers = ["Employé", "Type", "Début", "Fin", "Jours", "Statut"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    for i, row in enumerate(data, 2):
        nom = f"{row.get('prenom','')} {row.get('nom','')}".strip()
        vals = [nom, row.get('type_conge',''), str(row.get('date_debut','')), str(row.get('date_fin','')), row.get('nombre_jours',''), row.get('statut','')]
        for c, v in enumerate(vals, 1):
            ws.cell(row=i, column=c, value=v).border = thin
    for c in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(c)].width = 15
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# EXPORT ROUTES
@app.route('/export/presences/pdf')
@login_required
def export_presences_pdf():
    my_only = request.args.get('my') == '1'
    emp = get_current_employee() if my_only else None
    conn = get_db()
    cur = get_cursor(conn)
    q = "SELECT p.*, e.nom, e.prenom FROM presences p JOIN employes e ON p.employe_id = e.id "
    params = []
    if my_only:
        if emp:
            q += "WHERE p.employe_id = %s "
            params.append(emp['id'])
        else:
            q += "WHERE FALSE "
    else:
        scope_where, scope_params = department_scope_sql('e')
        q += f"WHERE {scope_where} "
        params.extend(scope_params)
    q += "ORDER BY p.date DESC LIMIT 500"
    cur.execute(q, params)
    data = cur.fetchall()
    cur.close(); conn.close()
    pdf = create_presences_pdf(data)
    resp = make_response(pdf.read())
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = 'attachment; filename=presences.pdf'
    return resp

@app.route('/export/presences/excel')
@login_required
def export_presences_excel():
    my_only = request.args.get('my') == '1'
    emp = get_current_employee() if my_only else None
    conn = get_db()
    cur = get_cursor(conn)
    q = "SELECT p.*, e.nom, e.prenom FROM presences p JOIN employes e ON p.employe_id = e.id "
    params = []
    if my_only:
        if emp:
            q += "WHERE p.employe_id = %s "
            params.append(emp['id'])
        else:
            q += "WHERE FALSE "
    else:
        scope_where, scope_params = department_scope_sql('e')
        q += f"WHERE {scope_where} "
        params.extend(scope_params)
    q += "ORDER BY p.date DESC LIMIT 800"
    cur.execute(q, params)
    data = cur.fetchall()
    cur.close(); conn.close()
    xlsx = create_presences_excel(data)
    resp = make_response(xlsx.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = 'attachment; filename=presences.xlsx'
    return resp

@app.route('/export/conges/pdf')
@login_required
def export_conges_pdf():
    my_only = request.args.get('my') == '1'
    emp = get_current_employee() if my_only else None
    conn = get_db()
    cur = get_cursor(conn)
    q = "SELECT c.*, e.nom, e.prenom FROM conges c JOIN employes e ON c.employe_id = e.id "
    params = []
    if my_only:
        if emp:
            q += "WHERE c.employe_id = %s "
            params.append(emp['id'])
        else:
            q += "WHERE FALSE "
    else:
        scope_where, scope_params = department_scope_sql('e')
        q += f"WHERE {scope_where} "
        params.extend(scope_params)
    q += "ORDER BY c.date_demande DESC LIMIT 500"
    cur.execute(q, params)
    data = cur.fetchall()
    cur.close(); conn.close()
    pdf = create_conges_pdf(data)
    resp = make_response(pdf.read())
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = 'attachment; filename=conges.pdf'
    return resp

@app.route('/export/conges/excel')
@login_required
def export_conges_excel():
    my_only = request.args.get('my') == '1'
    emp = get_current_employee() if my_only else None
    conn = get_db()
    cur = get_cursor(conn)
    q = "SELECT c.*, e.nom, e.prenom FROM conges c JOIN employes e ON c.employe_id = e.id "
    params = []
    if my_only:
        if emp:
            q += "WHERE c.employe_id = %s "
            params.append(emp['id'])
        else:
            q += "WHERE FALSE "
    else:
        scope_where, scope_params = department_scope_sql('e')
        q += f"WHERE {scope_where} "
        params.extend(scope_params)
    q += "ORDER BY c.date_demande DESC"
    cur.execute(q, params)
    data = cur.fetchall()
    cur.close(); conn.close()
    xlsx = create_conges_excel(data)
    resp = make_response(xlsx.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = 'attachment; filename=conges.xlsx'
    return resp

# ==================== BASIC ROUTES ====================
# Tableau de bord général extrait dans blueprints/dashboard.py.

# ==================== PRÉSENCES ==============================================
# Routes extraites dans blueprints/presences.py.

# Congés extraits dans blueprints/conges.py.


# ==================== ABSENCES NON JUSTIFIÉES ====================
# Jours ouvrés pris en compte pour la génération automatique des absences
# (0 = lundi ... 6 = dimanche). Par défaut : lundi → vendredi.
JOURS_OUVRES = {0, 1, 2, 3, 4}

# Fenêtre max de génération (en jours). Borné pour rester rapide et léger en
# mémoire sur les hébergements contraints (Render : timeout worker ~30 s,
# RAM 512 Mo). Ajustable.
LOOKBACK_JOURS = 365

def _dates_couvertes(cur, employe_id):
    """Ensemble des dates qui ne doivent PAS être comptées comme une absence
    pour un employé : jours avec présence + jours couverts par un congé
    approuvé + jours couverts par une permission approuvée."""
    couverts = set()

    # Présences enregistrées
    cur.execute("SELECT date FROM presences WHERE employe_id = %s", (employe_id,))
    for row in cur.fetchall():
        couverts.add(row['date'])

    # Congés approuvés (tous types de congé)
    cur.execute("""
        SELECT date_debut, date_fin FROM conges
        WHERE employe_id = %s AND statut = 'approuvé'
    """, (employe_id,))
    for row in cur.fetchall():
        d, fin = row.get('date_debut'), row.get('date_fin')
        if not d or not fin:
            continue  # plage incomplète (dates NULL) : on l'ignore
        while d <= fin:
            couverts.add(d)
            d += timedelta(days=1)

    # Permissions approuvées (module séparé, mais un jour de permission
    # approuvée n'est pas non plus une absence non justifiée)
    cur.execute("""
        SELECT date_debut, date_fin FROM permissions
        WHERE employe_id = %s AND statut = 'approuvé'
    """, (employe_id,))
    for row in cur.fetchall():
        d, fin = row.get('date_debut'), row.get('date_fin')
        if not d or not fin:
            continue
        while d <= fin:
            couverts.add(d)
            d += timedelta(days=1)

    return couverts


def generer_absences_automatiques(cur, date_jusqua=None, date_depuis=None,
                                  departement=None):
    """Enregistre automatiquement dans `absences` chaque jour ouvré passé SANS
    présence (et non couvert par un congé/permission approuvé) = « tout jour où
    l'employé n'a pas de présence ».

    Règles : jour ouvré (lun→ven) ; >= date d'embauche ; <= la veille ; aucune
    présence ; non couvert par un congé ou une permission approuvés.

    IMPORTANT (perf/mémoire) : la génération est BORNÉE à une fenêtre récente
    (LOOKBACK_JOURS) et les insertions sont envoyées par LOTS via execute_values
    (1 requête par lot au lieu d'1 par ligne). Sans cela, sur un hébergement
    contraint (Render : timeout worker 30 s, RAM 512 Mo), la génération depuis
    la date d'embauche + executemany => timeout worker + kill OOM => HTTP 500.
    Idempotent (UNIQUE(employe_id, date)). Retourne le nb d'absences créées.
    """
    if date_jusqua is None:
        date_jusqua = date.today() - timedelta(days=1)  # on exclut aujourd'hui
    if date_depuis is None:
        date_depuis = date_jusqua - timedelta(days=LOOKBACK_JOURS - 1)

    def _to_date(v):
        return v if isinstance(v, date) else datetime.strptime(str(v), '%Y-%m-%d').date()

    fin_globale = _to_date(date_jusqua)
    debut_global = _to_date(date_depuis)

    if departement is None:
        cur.execute("SELECT id, date_embauche FROM employes WHERE actif ORDER BY id")
    else:
        cur.execute("""SELECT id, date_embauche FROM employes
                       WHERE actif AND departement = %s ORDER BY id""", (departement,))
    employes = cur.fetchall()

    nb_creees = 0
    motif_auto = "Aucune présence enregistrée (généré automatiquement)"

    for emp in employes:
        embauche = emp['date_embauche']
        if not embauche:
            continue  # sans date d'embauche, période indéterminée
        debut = max(debut_global, _to_date(embauche))
        if debut > fin_globale:
            continue

        couverts = _dates_couvertes(cur, emp['id'])

        cur.execute(
            "SELECT date FROM absences WHERE employe_id = %s AND date BETWEEN %s AND %s",
            (emp['id'], debut, fin_globale),
        )
        deja = {row['date'] for row in cur.fetchall()}
        # Absences supprimées manuellement : ne JAMAIS les régénérer
        cur.execute("SELECT date FROM absences_exclues WHERE employe_id = %s", (emp['id'],))
        deja |= {row['date'] for row in cur.fetchall()}

        valeurs = []
        jour = debut
        while jour <= fin_globale:
            if jour.weekday() in JOURS_OUVRES and jour not in couverts and jour not in deja:
                valeurs.append((emp['id'], jour, motif_auto))
            jour += timedelta(days=1)

        if valeurs:
            execute_values(
                cur,
                "INSERT INTO absences (employe_id, date, motif) "
                "VALUES %s ON CONFLICT (employe_id, date) DO NOTHING",
                valeurs,
                page_size=500,
            )
            nb_creees += len(valeurs)

    return nb_creees


SEUIL_ALERTE_EXPIRATION_DOCUMENTS_JOURS = 30


def job_alertes_expiration_documents():
    """Job planifié (1x/jour) : alerte RH/admin (+ l'employé concerné) pour
    les documents dont la date d'expiration approche ou est dépassée
    (CDD, visa, certification, contrat...).

    Deux alertes possibles par document, chacune envoyée UNE SEULE FOIS
    (table `documents_alertes`) :
      - 'bientot' : expire dans <= 30 jours (mais pas encore expiré)
      - 'expire'  : date d'expiration dépassée
    """
    with app.app_context():
        try:
            with db_cursor(commit=True) as (conn, cur):
                cur.execute("""
                    INSERT INTO scheduler_runs (job_name, run_date)
                    VALUES ('alertes_expiration_documents', CURRENT_DATE)
                    ON CONFLICT DO NOTHING
                    RETURNING job_name
                """)
                if cur.fetchone() is None:
                    logger.info("Job alertes expiration documents : déjà exécuté aujourd'hui, on saute.")
                    return

                cur.execute("""
                    SELECT d.id, d.titre, d.date_expiration, d.employe_id,
                           e.nom, e.prenom, e.email
                    FROM documents d
                    LEFT JOIN employes e ON d.employe_id = e.id
                    WHERE d.date_expiration IS NOT NULL
                      AND d.date_expiration <= CURRENT_DATE + INTERVAL %s
                """, (f"{SEUIL_ALERTE_EXPIRATION_DOCUMENTS_JOURS} days",))
                candidats = cur.fetchall()
                if not candidats:
                    return

                cur.execute("SELECT id, employe_id FROM users WHERE employe_id IS NOT NULL")
                user_id_par_employe = {u['employe_id']: u['id'] for u in cur.fetchall()}
                cur.execute("SELECT id FROM users WHERE role IN ('admin', 'rh')")
                ids_rh = [u['id'] for u in cur.fetchall()]

                a_notifier = []
                for d in candidats:
                    type_alerte = 'expire' if d['date_expiration'] <= date.today() else 'bientot'
                    cur.execute("""
                        INSERT INTO documents_alertes (document_id, type_alerte)
                        VALUES (%s, %s) ON CONFLICT DO NOTHING
                        RETURNING document_id
                    """, (d['id'], type_alerte))
                    if cur.fetchone() is None:
                        continue  # déjà alerté pour ce document + ce type
                    a_notifier.append((d, type_alerte))

                if not a_notifier:
                    return

                for d, type_alerte in a_notifier:
                    proprietaire = f"{d['prenom']} {d['nom']}" if d['nom'] else "document général"
                    if type_alerte == 'expire':
                        message = f"« {d['titre']} » ({proprietaire}) a expiré le {d['date_expiration']}."
                    else:
                        message = f"« {d['titre']} » ({proprietaire}) expire le {d['date_expiration']}."

                    for uid in ids_rh:
                        create_notification(uid, "Document arrivant à expiration" if type_alerte == 'bientot'
                                             else "Document expiré", message, "warning")

                    if d['employe_id']:
                        uid = user_id_par_employe.get(d['employe_id'])
                        if uid:
                            create_notification(uid, "Votre document arrive à expiration" if type_alerte == 'bientot'
                                                 else "Votre document a expiré", message, "warning",
                                                 cur=cur)
                        queue_email(
                            d.get('email'),
                            "Votre document arrive à expiration" if type_alerte == 'bientot'
                            else "Votre document a expiré",
                            message + " Contactez les RH pour son renouvellement.",
                            cur=cur,
                            event_key=f"document-expiration:{d['id']}:{type_alerte}",
                        )

                details = "\n".join(
                    f"- [{'EXPIRÉ' if t == 'expire' else 'bientôt'}] {d['titre']} "
                    f"({d['prenom']} {d['nom']} — {d['date_expiration']})" if d['nom']
                    else f"- [{'EXPIRÉ' if t == 'expire' else 'bientôt'}] {d['titre']} ({d['date_expiration']})"
                    for d, t in a_notifier
                )
                _envoyer_email_texte(
                    [get_admin_email()],
                    f"📄 {len(a_notifier)} document(s) à vérifier (expiration)",
                    "Bonjour,\n\nLes documents suivants nécessitent votre attention :\n\n"
                    f"{details}\n\nConsultez la page Documents pour les renouveler si besoin.",
                    event_key=f"documents-expiration:{date.today().isoformat()}"
                )
        except Exception:
            logger.exception("Erreur lors du job planifié d'alertes d'expiration de documents")



def _envoyer_email_texte(destinataires, sujet, corps, event_key=None):
    """Met en file un e-mail texte ; aucun appel SMTP n'a lieu ici."""
    try:
        for index, destinataire in enumerate(destinataires or []):
            queue_email(destinataire, sujet, corps,
                        event_key=f"{event_key}:{index}" if event_key else None)
        return True
    except Exception as e:
        logger.error("Erreur mise en file e-mail texte: %s", e, exc_info=True)
        return False


def job_generation_quotidienne_absences():
    """Job planifié (1x/jour, voir `demarrer_scheduler`) : génère les absences
    automatiques (remplace l'ancienne génération au chargement de la page, qui
    empêchait les suppressions de "tenir"), puis notifie RH/admin ainsi que
    chaque employé concerné.

    Idempotent via `scheduler_runs` : si le job a déjà tourné aujourd'hui
    (redémarrage du service, plusieurs workers gunicorn...), il ne s'exécute
    pas deux fois et n'envoie donc pas deux fois les mêmes notifications/emails.
    """
    with app.app_context():
        try:
            with db_cursor(commit=True) as (conn, cur):
                cur.execute("""
                    INSERT INTO scheduler_runs (job_name, run_date)
                    VALUES ('generation_absences', CURRENT_DATE)
                    ON CONFLICT DO NOTHING
                    RETURNING job_name
                """)
                if cur.fetchone() is None:
                    logger.info("Job génération absences : déjà exécuté aujourd'hui, on saute.")
                    return

                nb = generer_absences_automatiques(cur)
                logger.info("Job génération absences : %d absence(s) créée(s).", nb)
                if nb == 0:
                    return

                motif_auto = "Aucune présence enregistrée (généré automatiquement)"
                cur.execute("""
                    SELECT a.employe_id, a.date, e.nom, e.prenom, e.email
                    FROM absences a JOIN employes e ON a.employe_id = e.id
                    WHERE a.date_enregistrement::date = CURRENT_DATE AND a.motif = %s
                    ORDER BY a.date
                """, (motif_auto,))
                nouvelles = cur.fetchall()

                # Notifier chaque employé concerné
                cur.execute("SELECT id, employe_id FROM users WHERE employe_id IS NOT NULL")
                user_id_par_employe = {u['employe_id']: u['id'] for u in cur.fetchall()}
                for a in nouvelles:
                    uid = user_id_par_employe.get(a['employe_id'])
                    message_absence = (
                        f"Aucune présence relevée le {a['date']} — une absence a été "
                        "enregistrée automatiquement. Vous pouvez déposer un justificatif "
                        "depuis votre espace employé."
                    )
                    if uid:
                        create_notification(uid, "Absence enregistrée", message_absence,
                                            "warning", cur=cur)
                    queue_email(
                        a.get('email'), "Absence enregistrée", message_absence,
                        cur=cur,
                        event_key=f"absence-auto:{a['employe_id']}:{a['date']}",
                    )

                # Notifier RH/admin (résumé global)
                cur.execute("SELECT id FROM users WHERE role IN ('admin', 'rh')")
                for u in cur.fetchall():
                    create_notification(
                        u['id'], "Absences générées automatiquement",
                        f"{nb} nouvelle(s) absence(s) détectée(s) (aucune présence enregistrée). "
                        f"Vérifiez la page Absences.",
                        "info"
                    )

                # Email récapitulatif à l'admin
                details = "\n".join(f"- {a['prenom']} {a['nom']} : {a['date']}" for a in nouvelles)
                _envoyer_email_texte(
                    [get_admin_email()],
                    f"🚫 {nb} absence(s) générée(s) automatiquement",
                    "Bonjour,\n\n"
                    f"{nb} absence(s) ont été enregistrée(s) automatiquement cette nuit "
                    "(aucune présence relevée le jour ouvré concerné) :\n\n"
                    f"{details}\n\n"
                    "Vous pouvez les consulter et les corriger si besoin sur la page Absences."
                )
        except Exception:
            logger.exception("Erreur lors du job planifié de génération des absences")


def job_purge_sessions():
    """Efface les sessions expirées ou révoquées de plus de 30 jours.

    Sans cela, le registre grossit indéfiniment et le compteur de sessions
    ouvertes finit par inclure des navigateurs fermés depuis longtemps.
    """
    try:
        with db_cursor(commit=True) as (conn, cur):
            cur.execute("""
                DELETE FROM sessions_actives
                 WHERE (revoked_at IS NOT NULL AND revoked_at < CURRENT_TIMESTAMP - INTERVAL '30 days')
                    OR (revoked_at IS NULL     AND last_seen  < CURRENT_TIMESTAMP - INTERVAL '30 days')
            """)
            n = cur.rowcount
        if n:
            logger.info("Purge des sessions : %s entrée(s) supprimée(s).", n)
    except Exception as e:
        logger.error("Erreur job_purge_sessions: %s", e, exc_info=True)


def job_validation_auto_maintenances():
    """Valide d'office les retours d'atelier restés sans réponse.

    Sans ce filet, une intervention resterait « en attente de validation »
    indéfiniment si le demandeur est absent, a quitté l'entreprise ou oublie
    simplement de répondre — et fausserait les indicateurs.
    """
    try:
        with db_cursor(commit=True) as (conn, cur):
            cur.execute("""
                SELECT mt.id, mt.exemplaire_id, ex.numero_inventaire
                  FROM materiel_maintenances mt
                  JOIN materiel_exemplaires ex ON ex.id = mt.exemplaire_id
                 WHERE mt.statut = 'a_valider'
                   AND mt.date_retour IS NOT NULL
                   AND mt.date_retour < CURRENT_DATE - make_interval(days => %s)
            """, (MAINTENANCE_VALIDATION_JOURS,))
            echues = cur.fetchall()
            for mt in echues:
                cur.execute("""UPDATE materiel_maintenances
                               SET statut = 'repare', valide_par = 'système',
                                   date_validation = CURRENT_DATE,
                                   validation_forcee = TRUE,
                                   cloture_par = 'système'
                               WHERE id = %s""", (mt['id'],))
        if echues:
            logger.info("Validation automatique de %s intervention(s) : %s",
                        len(echues), ", ".join(m['numero_inventaire'] for m in echues))
    except Exception as e:
        logger.error("Erreur job_validation_auto_maintenances: %s", e, exc_info=True)


def job_recalcul_soldes_conges():
    """Job planifié (1x/jour, voir `demarrer_scheduler`) : recalcule le nombre
    de jours de congé acquis (`jours_acquis`) de chaque employé pour l'année en
    cours, selon l'accumulation mensuelle (~2,08 j/mois, proratisée pour les
    nouveaux embauchés). Ne touche jamais les soldes fixés manuellement par
    RH/admin (`jours_acquis_manuel = TRUE`).

    Tourner ce job une fois par jour suffit très largement (le résultat ne
    change qu'au changement de mois), mais c'est sans risque de le faire
    tourner plus souvent : le calcul est idempotent.
    """
    with app.app_context():
        try:
            with db_cursor(commit=True) as (conn, cur):
                annee_courante = datetime.now().year
                cur.execute("""
                    SELECT e.id, e.date_embauche
                    FROM employes e
                    LEFT JOIN soldes_conges s ON s.employe_id = e.id AND s.annee = %s
                    WHERE COALESCE(s.jours_acquis_manuel, FALSE) = FALSE
                """, (annee_courante,))
                employes = cur.fetchall()

                nb_mis_a_jour = 0
                for emp in employes:
                    nouveau_solde = calculer_jours_acquis_prorata(emp.get('date_embauche'), annee_courante)
                    cur.execute("""
                        INSERT INTO soldes_conges (employe_id, annee, jours_acquis, jours_utilises)
                        VALUES (%s, %s, %s, 0)
                        ON CONFLICT (employe_id, annee) DO UPDATE
                        SET jours_acquis = %s
                        WHERE soldes_conges.jours_acquis_manuel = FALSE
                    """, (emp['id'], annee_courante, nouveau_solde, nouveau_solde))
                    nb_mis_a_jour += 1
                logger.info("Job recalcul soldes congés : %d employé(s) mis à jour (accumulation mensuelle).",
                            nb_mis_a_jour)
        except Exception:
            logger.exception("Erreur lors du job planifié de recalcul des soldes de congés")


def job_traiter_file_emails():
    """Vide périodiquement l'outbox SMTP avec reprise et tentatives bornées."""
    if not app.config.get('EMAIL_ENABLED'):
        return {'traites': 0, 'envoyes': 0, 'replanifies': 0, 'echecs': 0}
    with app.app_context():
        resultat = traiter_outbox(
            db_cursor, _send_outbox_message, logger,
            taille_lot=app.config.get('EMAIL_BATCH_SIZE', 20),
            tentatives_max=app.config.get('EMAIL_MAX_ATTEMPTS', 5),
        )
        if resultat['traites']:
            logger.info("Outbox e-mail : %s", resultat)
        return resultat


def demarrer_scheduler():
    """Repli local optionnel ; la production utilise ``scheduler_worker.py``.

    ``SCHEDULER_MODE=embedded`` conserve l'expérience historique pour un poste
    de développement mono-processus. Le mode par défaut est ``disabled`` en
    production et ``embedded`` ailleurs. Gunicorn ne lance donc plus de tâches.
    """
    mode_defaut = 'disabled' if os.environ.get('FLASK_ENV') == 'production' else 'embedded'
    mode = os.environ.get('SCHEDULER_MODE', mode_defaut).lower()
    if os.environ.get('FLASK_ENV') == 'testing' or mode != 'embedded':
        return None
    debug_actif = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    if debug_actif and os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
        return None

    from services.scheduler_runtime import build_scheduler
    import atexit

    scheduler = build_scheduler(
        jobs={
            'generation_absences': job_generation_quotidienne_absences,
            'alertes_documents': job_alertes_expiration_documents,
            'recalcul_soldes': job_recalcul_soldes_conges,
            'alertes_contrats': job_alertes_contrats,
            'purge_sessions': job_purge_sessions,
            'validation_maintenances': job_validation_auto_maintenances,
            'email_outbox': job_traiter_file_emails,
        },
        db_cursor=db_cursor,
        app_config=app.config,
        blocking=False,
    )
    scheduler.phase4_heartbeat()
    scheduler.start()
    atexit.register(lambda: scheduler.shutdown(wait=False))
    logger.warning(
        "Scheduler embarqué actif pour le développement. En production, "
        "utilisez python scheduler_worker.py."
    )
    return scheduler


# Routes absences extraites dans blueprints/absences.py.


# ==================== PERMISSIONS (MODULE SÉPARÉ, COMME LES CONGÉS) ====================
# Une permission est demandée / approuvée / refusée exactement comme un congé,
# mais elle vit dans sa propre table (`permissions`) et n'a AUCUN impact sur le
# solde de congés (`soldes_conges`).
@app.route('/permissions')
@login_required
def permissions():
    search = request.args.get('search', '').strip()
    statut = request.args.get('statut', '').strip()
    date_debut = request.args.get('date_debut', '').strip()
    date_fin = request.args.get('date_fin', '').strip()
    page = max(1, request.args.get('page', 1, type=int))
    per_page = 10

    where = ""
    params = []
    if search:
        where += " AND (LOWER(e.nom) LIKE %s OR LOWER(e.prenom) LIKE %s)"
        params += [f"%{search.lower()}%", f"%{search.lower()}%"]
    if statut:
        where += " AND p.statut = %s"; params.append(statut)
    if date_debut:
        where += " AND p.date_debut >= %s"; params.append(date_debut)
    if date_fin:
        where += " AND p.date_fin <= %s"; params.append(date_fin)

    with db_cursor() as (conn, cur):
        scope_where, scope_params = department_scope_sql('e', cur=cur)
        where += f" AND {scope_where}"
        params += scope_params
        from_ = "permissions p JOIN employes e ON p.employe_id = e.id"
        cur.execute(f"SELECT COUNT(*) AS nb FROM {from_} WHERE 1=1{where}", params)
        total = cur.fetchone()['nb']
        pg = pagination_info(total, page, per_page)
        offset = (pg['page'] - 1) * per_page
        cur.execute(f"SELECT p.*, e.nom, e.prenom FROM {from_} WHERE 1=1{where} ORDER BY p.date_demande DESC LIMIT %s OFFSET %s", params + [per_page, offset])
        permissions_list = cur.fetchall()
        cur.execute(f"SELECT id, nom, prenom FROM employes e WHERE {scope_where} ORDER BY nom",
                    scope_params)
        employees = cur.fetchall()
    filters = {'search': search, 'statut': statut, 'date_debut': date_debut, 'date_fin': date_fin}
    return render_template('permissions.html', permissions=permissions_list, employees=employees, filters=filters,
                           libelles=DEMANDE_LIBELLES, ouvertes=DEMANDE_OUVERTES,
                           peut_decider=_peut_decider_rh(), peut_avis=_peut_donner_avis(),
                           pg=pg, page_items=page_list(pg['page'], pg['pages']),
                           base_qs=urlencode({k: v for k, v in filters.items() if v}))


@app.route('/permissions/add', methods=['GET', 'POST'])
@login_required
def add_permission():
    """Dépôt d'une demande de permission (même circuit que les congés)."""
    moi = get_current_employee()
    gestionnaire = session.get('role') in ('admin', 'rh', 'manager')

    with db_cursor(commit=True) as (conn, cur):
        scope_where, scope_params = department_scope_sql('e', cur=cur)
        if request.method == 'POST':
            employe_id = request.form.get('employe_id', type=int)
            date_debut = request.form.get('date_debut')
            date_fin = request.form.get('date_fin')
            motif = request.form.get('motif', '').strip()

            if not gestionnaire:
                if not moi:
                    flash("Aucun employé n'est lié à votre compte : "
                          "contactez les RH.", "warning")
                    return redirect(url_for('self_service'))
                employe_id = moi['id']

            if employe_id and date_debut and date_fin:
                d1 = datetime.strptime(date_debut, '%Y-%m-%d')
                d2 = datetime.strptime(date_fin, '%Y-%m-%d')
                if d2 < d1:
                    flash("La date de fin ne peut pas être avant la date de début", "danger")
                    cur.execute(f"SELECT id, nom, prenom FROM employes e WHERE {scope_where} ORDER BY nom", scope_params)
                    return render_template('permission_form.html', employees=cur.fetchall(),
                                           moi=moi, gestionnaire=gestionnaire)
                nombre_jours = (d2 - d1).days + 1
                cur.execute("""
                    INSERT INTO permissions (employe_id, motif, date_debut, date_fin,
                                             nombre_jours, statut, demande_par_id)
                    VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id
                """, (employe_id, motif, date_debut, date_fin, nombre_jours,
                      DEMANDE_EN_ATTENTE, session.get('user_id')))
                pid = cur.fetchone()['id']

                nom, dept = _libelle_employe(cur, employe_id)
                titre = "Demande de permission : %s" % nom
                corps = ("%s jour(s) du %s au %s. Votre avis est attendu."
                         % (nombre_jours, d1.strftime('%d/%m/%Y'), d2.strftime('%d/%m/%Y')))
                managers = _managers_du_departement(cur, dept)
                for m in managers:
                    create_notification(m['id'], titre, corps, 'info', cur=cur)
                    queue_email(m.get('email'), titre, corps, cur=cur,
                                event_key=f"permission-a-traiter:{pid}:{m['id']}")
                if not managers:
                    message_rh = corps + " (aucun manager sur ce département)"
                    _notifier_roles(cur, ('admin', 'rh'), titre, message_rh,
                                    'info', sauf=session.get('user_id'))
                    _envoyer_roles(cur, ('admin', 'rh'), titre, message_rh,
                                   f"permission-a-traiter:{pid}")

                log_action(session.get('user_id'), session.get('username'),
                           "Demande de permission", "permission", pid, f"{nombre_jours} j")
                flash("Demande de permission soumise. Vous serez notifié à chaque étape.",
                      "success")
                return redirect(url_for('self_service_conges') if not gestionnaire
                                else url_for('permissions'))
            flash("Veuillez remplir tous les champs obligatoires", "danger")
        cur.execute(f"SELECT id, nom, prenom FROM employes e WHERE {scope_where} ORDER BY nom", scope_params)
        employees = cur.fetchall()
    return render_template('permission_form.html', employees=employees,
                           moi=moi, gestionnaire=gestionnaire)


@app.route('/permissions/avis/<int:id>', methods=['POST'])
@login_required
@role_required('admin', 'rh', 'manager')
def avis_permission(id):
    """Étape 2 : avis du manager du département."""
    avis = (request.form.get('avis') or '').strip()
    commentaire = (request.form.get('commentaire') or '').strip()
    if avis not in ('favorable', 'defavorable'):
        flash("Avis invalide.", "danger")
        return redirect(url_for('permissions'))
    if avis == 'defavorable' and not commentaire:
        flash("Merci de motiver un avis défavorable.", "danger")
        return redirect(url_for('permissions'))

    with db_cursor(commit=True) as (conn, cur):
        cur.execute("SELECT * FROM permissions WHERE id = %s", (id,))
        pm = cur.fetchone()
        if not pm:
            flash("Demande introuvable.", "danger")
            return redirect(url_for('permissions'))
        if pm['statut'] not in DEMANDE_OUVERTES:
            flash("Cette demande est déjà tranchée.", "warning")
            return redirect(url_for('permissions'))

        cur.execute("""UPDATE permissions SET statut = %s, avis_manager = %s,
                          avis_manager_par = %s, avis_manager_le = CURRENT_DATE,
                          avis_commentaire = %s
                       WHERE id = %s""",
                    (DEMANDE_AVIS_RENDU, avis, session.get('username'),
                     commentaire or None, id))

        nom, _ = _libelle_employe(cur, pm['employe_id'])
        sujet_rh = "Permission à décider : %s" % nom
        message_rh = ("Avis %s du manager %s. %s"
                      % (avis, session.get('username'), commentaire[:120]))
        _notifier_roles(cur, ('admin', 'rh'), sujet_rh, message_rh,
                        'info', sauf=session.get('user_id'))
        _envoyer_roles(cur, ('admin', 'rh'), sujet_rh, message_rh,
                       f"permission-a-decider:{id}")
        uid = _user_id_de_employe(cur, pm['employe_id'])
        if uid:
            create_notification(uid, "Votre demande de permission avance",
                                "Votre manager a rendu un avis %s. "
                                "La décision RH suit." % avis, 'info')

    log_action(session.get('user_id'), session.get('username'),
               "Avis manager permission", "permission", id, avis)
    flash("Avis enregistré : les RH vont trancher.", "success")
    return redirect(url_for('permissions'))


@app.route('/permissions/update/<int:id>', methods=['POST'])
@login_required
@role_required('admin', 'rh')
def update_permission(id):
    """Étape 3 : décision finale des RH."""
    action = request.form.get('action')
    motif_refus = (request.form.get('motif_refus') or '').strip()
    if action == 'refuser' and not motif_refus:
        flash("Merci d'indiquer le motif du refus : l'employé en sera informé.", "danger")
        return redirect(url_for('permissions'))

    with db_cursor(commit=True) as (conn, cur):
        cur.execute("SELECT * FROM permissions WHERE id = %s", (id,))
        pm = cur.fetchone()
        if not pm:
            flash("Demande introuvable.", "danger")
            return redirect(url_for('permissions'))
        if pm['statut'] not in DEMANDE_OUVERTES:
            flash("Cette demande est déjà tranchée.", "warning")
            return redirect(url_for('permissions'))

        if action == 'approuver':
            cur.execute("""UPDATE permissions SET statut = %s, decide_par = %s,
                              decide_le = CURRENT_DATE, motif_refus = NULL
                           WHERE id = %s""",
                        (DEMANDE_APPROUVEE, session.get('username'), id))
            _notifier_employe_evenement(
                cur, pm['employe_id'], "Permission approuvée",
                "Votre permission du %s au %s est approuvée."
                % (pm['date_debut'], pm['date_fin']), 'success',
                cle_evenement=f"permission-decision:{id}:approuve",
            )
            flash("Permission approuvée", "success")
        elif action == 'refuser':
            cur.execute("""UPDATE permissions SET statut = %s, decide_par = %s,
                              decide_le = CURRENT_DATE, motif_refus = %s
                           WHERE id = %s""",
                        (DEMANDE_REFUSEE, session.get('username'), motif_refus, id))
            _notifier_employe_evenement(
                cur, pm['employe_id'], "Permission refusée",
                "Votre demande du %s au %s a été refusée : %s"
                % (pm['date_debut'], pm['date_fin'], motif_refus), 'danger',
                cle_evenement=f"permission-decision:{id}:refuse",
            )
            flash("Permission refusée : l'employé est informé du motif.", "info")
        else:
            flash("Action inconnue.", "danger")
            return redirect(url_for('permissions'))

    log_action(session.get('user_id'), session.get('username'),
               "Décision permission", "permission", id, action)
    return redirect(url_for('permissions'))


@app.route('/permissions/<int:id>/annuler', methods=['POST'])
@login_required
def annuler_permission(id):
    """Retrait d'une demande de permission par son auteur."""
    with db_cursor(commit=True) as (conn, cur):
        cur.execute("SELECT * FROM permissions WHERE id = %s", (id,))
        pm = cur.fetchone()
        if not pm:
            flash("Demande introuvable.", "danger")
            return redirect(url_for('self_service_conges'))
        moi = get_current_employee()
        est_le_mien = moi and moi['id'] == pm['employe_id']
        if not (est_le_mien or _peut_decider_rh()):
            flash("Vous ne pouvez annuler que vos propres demandes.", "danger")
            return redirect(url_for('self_service_conges'))
        if pm['statut'] not in DEMANDE_OUVERTES:
            flash("Cette demande est déjà tranchée : elle ne peut plus être annulée.",
                  "warning")
            return redirect(url_for('self_service_conges') if est_le_mien
                            else url_for('permissions'))

        cur.execute("""UPDATE permissions SET statut = %s, annule_par = %s
                       WHERE id = %s""",
                    (DEMANDE_ANNULEE, session.get('username'), id))
        nom, dept = _libelle_employe(cur, pm['employe_id'])
        for m in _managers_du_departement(cur, dept):
            create_notification(m['id'], "Demande de permission annulée",
                                "%s a retiré sa demande du %s au %s."
                                % (nom, pm['date_debut'], pm['date_fin']), 'info')

    log_action(session.get('user_id'), session.get('username'),
               "Annulation permission", "permission", id, None)
    flash("Demande annulée.", "success")
    return redirect(url_for('self_service_conges') if est_le_mien
                    else url_for('permissions'))


@app.route('/permissions/delete/<int:id>', methods=['POST'])
@login_required
@role_required('admin', 'rh', 'manager')
def delete_permission(id):
    with db_cursor(commit=True) as (conn, cur):
        cur.execute("DELETE FROM permissions WHERE id = %s", (id,))
    flash("Demande de permission supprimée", "success")
    return redirect(url_for('permissions'))


@app.route('/soldes-conges')
@login_required
@role_required('admin', 'rh', 'manager')
def soldes_conges_page():
    """Affiche le solde de congés de chaque employé pour une année donnée.
    jours_utilises est recalculé automatiquement depuis les congés approuvés
    (via recalculer_solde) avant affichage, donc toujours à jour."""
    annee = request.args.get('annee', type=int) or datetime.now().year

    with db_cursor(commit=True) as (conn, cur):
        scope_where, scope_params = department_scope_sql('e', cur=cur)
        cur.execute(f"""SELECT id, nom, prenom FROM employes e
                        WHERE {scope_where} ORDER BY nom, prenom""", scope_params)
        employees = cur.fetchall()
        for emp in employees:
            recalculer_solde(emp['id'], annee, cur=cur)

    soldes = []
    for emp in employees:
        s = get_solde_conges(emp['id'], annee)
        s['employe_id'] = emp['id']
        s['nom'] = emp['nom']
        s['prenom'] = emp['prenom']
        soldes.append(s)

    annee_courante = datetime.now().year
    annees_disponibles = list(range(annee_courante - 2, annee_courante + 2))

    return render_template('soldes_conges.html', soldes=soldes, annee=annee,
                            annees_disponibles=annees_disponibles)


@app.route('/soldes-conges/update/<int:employe_id>', methods=['POST'])
@login_required
@role_required('admin', 'rh')
def update_solde_conges(employe_id):
    """Permet de modifier manuellement le nombre de jours acquis (allocation)
    d'un employé pour une année. jours_utilises reste calculé automatiquement."""
    annee = request.form.get('annee', type=int) or datetime.now().year
    jours_acquis = request.form.get('jours_acquis', type=float)

    if jours_acquis is None or jours_acquis < 0:
        flash("Valeur de jours acquis invalide", "danger")
        return redirect(url_for('soldes_conges_page', annee=annee))

    with db_cursor(commit=True) as (conn, cur):
        cur.execute("""
            INSERT INTO soldes_conges (employe_id, annee, jours_acquis, jours_utilises, jours_acquis_manuel)
            VALUES (%s, %s, %s, 0, TRUE)
            ON CONFLICT (employe_id, annee)
            DO UPDATE SET jours_acquis = %s, jours_acquis_manuel = TRUE
        """, (employe_id, annee, jours_acquis, jours_acquis))
        cur.execute("SELECT nom, prenom FROM employes WHERE id = %s", (employe_id,))
        emp = cur.fetchone()

    if emp:
        log_action(session.get('user_id'), session.get('username'), "UPDATE_SOLDE_CONGES",
                   "solde_conges", employe_id, f"{emp['prenom']} {emp['nom']} → jours_acquis={jours_acquis} ({annee})")

    flash("Solde de congés mis à jour avec succès", "success")
    return redirect(url_for('soldes_conges_page', annee=annee))


@app.route('/audit')
@role_required('admin', 'rh')
def audit():
    with db_cursor() as (conn, cur):
        cur.execute("SELECT a.*, u.username FROM audit_logs a LEFT JOIN users u ON a.user_id = u.id ORDER BY a.timestamp DESC LIMIT 150")
        logs = cur.fetchall()
    return render_template('audit.html', logs=logs)



@app.route('/employes')
@login_required
def index():
    conn = get_db()
    cur = get_cursor(conn)
    scope_where, scope_params = department_scope_sql('e', cur=cur)

    search = request.args.get('search', '').strip()
    selected_dept = request.args.get('departement', '').strip()
    sort = request.args.get('sort', 'nom')
    order = request.args.get('order', 'asc')
    page = max(1, request.args.get('page', 1, type=int))
    per_page = 10

    where = f" AND {scope_where}"
    params = list(scope_params)
    if search:
        where += " AND (LOWER(e.nom) LIKE %s OR LOWER(e.prenom) LIKE %s OR LOWER(e.poste) LIKE %s OR LOWER(e.email) LIKE %s)"
        s = f"%{search.lower()}%"
        params += [s, s, s, s]
    if selected_dept:
        where += " AND e.departement = %s"; params.append(selected_dept)

    sort_map = {'nom': 'e.nom, e.prenom', 'salaire': 'COALESCE(e.salaire, 0)', 'date_embauche': 'e.date_embauche', 'poste': 'e.poste'}
    sort_col = sort_map.get(sort, 'nom, prenom')
    direction = 'DESC' if order.lower() == 'desc' else 'ASC'
    order_clause = f" ORDER BY {sort_col} {direction}"

    cur.execute(f"SELECT COUNT(*) AS nb FROM employes e WHERE 1=1{where}", params)
    total = cur.fetchone()['nb']
    pg = pagination_info(total, page, per_page)
    offset = (pg['page'] - 1) * per_page
    # La photo est portée par le compte (users.photo), pas par la fiche employé.
    # Sous-requête scalaire plutôt que JOIN : un même employé peut avoir
    # plusieurs comptes, une jointure dupliquerait la ligne. On privilégie le
    # compte qui a effectivement une photo.
    cur.execute(f"""SELECT e.*, (
                        SELECT u.photo FROM users u
                         WHERE u.employe_id = e.id AND u.photo IS NOT NULL
                         ORDER BY u.id LIMIT 1
                    ) AS photo
                    FROM employes e WHERE 1=1{where}{order_clause} LIMIT %s OFFSET %s""",
                params + [per_page, offset])
    employes = cur.fetchall()

    for emp in employes:
        cur.execute("""
            SELECT date, heure_arrivee, statut
            FROM presences
            WHERE employe_id = %s
            ORDER BY date DESC
            LIMIT 1
        """, (emp['id'],))
        last = cur.fetchone()
        if last:
            emp['last_presence'] = dict(last)
            if emp['last_presence'].get('heure_arrivee'):
                emp['last_presence']['heure_arrivee'] = str(emp['last_presence']['heure_arrivee'])[:5]
        else:
            emp['last_presence'] = None

    dept_where, dept_params = department_scope_sql('d', 'nom', cur)
    cur.execute(f"SELECT DISTINCT d.nom FROM departements d WHERE {dept_where} ORDER BY d.nom",
                dept_params)
    depts = cur.fetchall()

    cur.execute(f"""
        SELECT COUNT(*) as total, COALESCE(AVG(e.salaire), 0) as salaire_moyen,
               COUNT(DISTINCT e.departement) as nb_departements
        FROM employes e WHERE {scope_where}
    """, scope_params)
    stats = cur.fetchone()
    stats = dict(stats) if stats else {'total': 0, 'salaire_moyen': 0, 'nb_departements': 0}

    cur.close()
    conn.close()
    filters = {'search': search, 'departement': selected_dept, 'sort': sort, 'order': order}
    return render_template('index.html', employes=employes, depts=depts, search=search, selected_dept=selected_dept,
                           sort=sort, order=order, stats=stats, pg=pg, page_items=page_list(pg['page'], pg['pages']),
                           base_qs=urlencode({k: v for k, v in filters.items() if v}))
@app.route('/employes/<int:id>')
@login_required
def view_employee(id):
    conn = get_db()
    cur = get_cursor(conn)
    # `photo` vient du compte lié (voir la remarque dans index()).
    cur.execute("""SELECT e.*, (
                       SELECT u.photo FROM users u
                        WHERE u.employe_id = e.id AND u.photo IS NOT NULL
                        ORDER BY u.id LIMIT 1
                   ) AS photo
                   FROM employes e WHERE e.id = %s""", (id,))
    employee = cur.fetchone()
    if not employee:
        cur.close()
        conn.close()
        flash("Employé introuvable", "danger")
        return redirect(url_for('index'))

    cur.execute("SELECT * FROM documents WHERE employe_id = %s ORDER BY date_upload DESC", (id,))
    employee_documents = cur.fetchall()

    # Dernière présence enregistrée (n'était auparavant jamais transmise au
    # template : la carte "Dernière présence" affichait donc toujours "aucune
    # présence enregistrée", même quand des présences existaient).
    cur.execute("""
        SELECT date, heure_arrivee, heure_depart, statut
        FROM presences WHERE employe_id = %s ORDER BY date DESC LIMIT 1
    """, (id,))
    last = cur.fetchone()
    last_presence = None
    if last:
        last_presence = dict(last)
        if last_presence.get('heure_arrivee'):
            last_presence['heure_arrivee'] = str(last_presence['heure_arrivee'])[:5]
        if last_presence.get('heure_depart'):
            last_presence['heure_depart'] = str(last_presence['heure_depart'])[:5]

    # Statistiques présence / absence / congés de l'année en cours
    annee_courante = date.today().year
    cur.execute("""
        SELECT COUNT(*) AS nb FROM presences
        WHERE employe_id = %s AND EXTRACT(YEAR FROM date) = %s
    """, (id, annee_courante))
    nb_presences = cur.fetchone()['nb']

    cur.execute("""
        SELECT heure_arrivee FROM presences
        WHERE employe_id = %s AND EXTRACT(YEAR FROM date) = %s AND heure_arrivee IS NOT NULL
    """, (id, annee_courante))
    nb_retards = sum(1 for r in cur.fetchall() if calculer_retard(r['heure_arrivee']) > 0)

    cur.execute("""
        SELECT COUNT(*) AS nb FROM absences
        WHERE employe_id = %s AND EXTRACT(YEAR FROM date) = %s
          AND COALESCE(statut, 'non_justifiee') <> 'acceptee'
    """, (id, annee_courante))
    nb_absences = cur.fetchone()['nb']

    cur.execute("""
        SELECT COALESCE(SUM(nombre_jours), 0) AS nb FROM conges
        WHERE employe_id = %s AND statut = 'approuvé' AND EXTRACT(YEAR FROM date_debut) = %s
    """, (id, annee_courante))
    nb_jours_conges = cur.fetchone()['nb']

    stats_presence = {
        'annee': annee_courante,
        'presences': nb_presences,
        'retards': nb_retards,
        'absences': nb_absences,
        'jours_conges': nb_jours_conges,
    }

    cur.close()
    conn.close()
    return render_template('detail.html', employee=employee, documents=employee_documents,
                           last_presence=last_presence, stats_presence=stats_presence,
                           today=date.today(),
                           bientot=date.today() + timedelta(days=SEUIL_ALERTE_EXPIRATION_DOCUMENTS_JOURS))

@app.route('/employes/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'rh')
def edit_employee(id):
    conn = get_db()
    cur = get_cursor(conn)

    if request.method == 'POST':
        nom = request.form.get('nom')
        prenom = request.form.get('prenom')
        poste = request.form.get('poste')
        departement = request.form.get('departement')
        email = request.form.get('email')
        telephone = request.form.get('telephone')
        salaire = request.form.get('salaire')
        
        cur.execute("""
            UPDATE employes 
            SET nom=%s, prenom=%s, poste=%s, departement=%s, email=%s, telephone=%s, salaire=%s
            WHERE id = %s
        """, (nom, prenom, poste, departement, email, telephone, salaire, id))
        conn.commit()
        flash("Employé modifié avec succès", "success")
        cur.close()
        conn.close()
        return redirect(url_for('index'))

    cur.execute("SELECT * FROM employes WHERE id = %s", (id,))
    employee = cur.fetchone()
    cur.execute("SELECT DISTINCT nom FROM departements ORDER BY nom")
    depts = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('form.html', employee=employee, depts=depts, title="Modifier l'employé")

@app.route('/employes/<int:id>/delete', methods=['POST'])
@login_required
@role_required('admin', 'rh')
def delete_employee(id):
    conn = get_db()
    cur = get_cursor(conn)
    
    try:
        # 1. Supprimer les présences liées
        cur.execute("DELETE FROM presences WHERE employe_id = %s", (id,))
        
        # 2. Supprimer les congés liés
        cur.execute("DELETE FROM conges WHERE employe_id = %s", (id,))

        # 2b. Supprimer les permissions liées (module séparé)
        cur.execute("DELETE FROM permissions WHERE employe_id = %s", (id,))

        # 2c. Supprimer les absences liées
        cur.execute("DELETE FROM absences WHERE employe_id = %s", (id,))
        
        # 3. Supprimer les soldes de congés liés
        cur.execute("DELETE FROM soldes_conges WHERE employe_id = %s", (id,))
        
        # 4. Supprimer les documents liés
        cur.execute("DELETE FROM documents WHERE employe_id = %s", (id,))
        
        # 5. Supprimer les notifications liées (si la table existe)
        try:
            cur.execute("DELETE FROM notifications WHERE user_id IN (SELECT id FROM users WHERE employe_id = %s)", (id,))
        except:
            pass
        
        # 6. Supprimer les utilisateurs liés (clé étrangère principale)
        cur.execute("DELETE FROM users WHERE employe_id = %s", (id,))
        
        # 7. Enfin supprimer l'employé
        cur.execute("DELETE FROM employes WHERE id = %s", (id,))
        
        conn.commit()
        flash("Employé et toutes ses données associées ont été supprimés avec succès", "success")
        
    except Exception as e:
        conn.rollback()
        flash(f"Erreur lors de la suppression : {str(e)}", "danger")
        logger.error("Erreur delete_employee: %s", e, exc_info=True)
    finally:
        cur.close()
        conn.close()
    
    return redirect(url_for('index'))

# ==================== RAPPORTS AVANCÉS ====================
@app.route('/rapports')
@login_required
def rapports():
    conn = get_db()
    cur = get_cursor(conn)
    scope_where, scope_params = department_scope_sql('e', cur=cur)
    
    # Filters
    date_debut = request.args.get('date_debut', '')
    date_fin = request.args.get('date_fin', '')
    employe_id = request.args.get('employe_id', '')
    type_rapport = request.args.get('type', 'presences')
    statut = request.args.get('statut', '')
    
    cur.execute(f"SELECT id, prenom, nom FROM employes e WHERE {scope_where} ORDER BY nom, prenom",
                scope_params)
    employees = cur.fetchall()
    
    presences_data = []
    conges_data = []
    total_jours = 0
    
    if type_rapport == 'presences':
        q = f"""SELECT p.*, e.nom, e.prenom FROM presences p
               JOIN employes e ON p.employe_id = e.id WHERE {scope_where} """
        params = list(scope_params)
        if date_debut:
            q += " AND p.date >= %s"
            params.append(date_debut)
        if date_fin:
            q += " AND p.date <= %s"
            params.append(date_fin)
        if employe_id:
            q += " AND p.employe_id = %s"
            params.append(int(employe_id))
        q += " ORDER BY p.date DESC LIMIT 200"
        cur.execute(q, params)
        presences_data = cur.fetchall()
        for p in presences_data:
            p['retard_minutes'] = calculer_retard(p['heure_arrivee'])
    else:
        q = f"""SELECT c.*, e.nom, e.prenom FROM conges c
               JOIN employes e ON c.employe_id = e.id WHERE {scope_where} """
        params = list(scope_params)
        if date_debut:
            q += " AND c.date_debut >= %s"
            params.append(date_debut)
        if date_fin:
            q += " AND c.date_fin <= %s"
            params.append(date_fin)
        if employe_id:
            q += " AND c.employe_id = %s"
            params.append(int(employe_id))
        if statut:
            q += " AND c.statut = %s"
            params.append(statut)
        q += " ORDER BY c.date_debut DESC LIMIT 200"
        cur.execute(q, params)
        conges_data = cur.fetchall()
        total_jours = sum((c['nombre_jours'] or 0) for c in conges_data)
    
    cur.close(); conn.close()
    
    return render_template('rapports.html', 
                           employees=employees,
                           presences=presences_data,
                           conges=conges_data,
                           date_debut=date_debut, date_fin=date_fin,
                           selected_employe=employe_id,
                           type_rapport=type_rapport,
                           statut=statut,
                           total_jours=total_jours)

# ==================== DOCUMENTS ===============================================
# Routes extraites dans blueprints/documents.py (Blueprint ``documents``).

# ==================== MAIN ====================

# ==================== STUB ROUTES (pour compatibilité templates) ====================
# ==================== HISTORIQUE DES PRÉSENCES ==============================
# Route extraite dans blueprints/presences.py.

# ==================== DÉPARTEMENTS ===========================================
# Routes extraites dans blueprints/departements.py.

# ==================== UTILISATEURS ET ACCÈS ==================================
# Routes extraites dans blueprints/utilisateurs.py.

@app.route('/add_employee', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'rh')
def add_employee():
    conn = get_db()
    cur = get_cursor(conn)

    if request.method == 'POST':
        nom = request.form.get('nom')
        prenom = request.form.get('prenom')
        poste = request.form.get('poste')
        departement = request.form.get('departement')
        email = request.form.get('email')
        telephone = request.form.get('telephone')
        salaire = request.form.get('salaire')
        date_embauche = request.form.get('date_embauche')
        
        if nom and prenom and poste:
            cur.execute("""
                INSERT INTO employes (nom, prenom, poste, departement, email, telephone, salaire, date_embauche)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (nom, prenom, poste, departement, email, telephone, salaire, date_embauche))
            employe_id = cur.fetchone()['id']
            libelle = f"{prenom} {nom}"
            _notifier_roles(
                cur, ('admin', 'rh'), "Nouvel employé enregistré",
                f"{libelle} a été ajouté au département {departement or 'non renseigné'}.",
                'success', sauf=session.get('user_id'))
            queue_email(
                email, "Bienvenue — dossier salarié créé",
                f"Bonjour {prenom},\n\nVotre dossier salarié a été créé dans Gestion du Personnel. "
                "Les RH vous communiqueront vos accès à l'espace employé.",
                cur=cur, event_key=f"employe-cree:{employe_id}")
            conn.commit()
            flash("Employé ajouté avec succès ; les personnes concernées ont été informées", "success")
            cur.close()
            conn.close()
            return redirect(url_for('index'))
        else:
            flash("Veuillez remplir les champs obligatoires", "danger")

    cur.execute("SELECT DISTINCT nom FROM departements ORDER BY nom")
    depts = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('form.html', employee=None, depts=depts, title="Nouvel employé")

@app.route('/calendrier-conges')
@login_required
def calendrier_conges():
    annee = request.args.get('annee', type=int) or datetime.now().year

    conn = get_db()
    cur = get_cursor(conn)
    scope_where, scope_params = department_scope_sql('e', cur=cur)
    cur.execute(f"""
        SELECT c.date_debut, c.date_fin, c.nombre_jours, c.statut,
               e.prenom, e.nom
        FROM conges c
        JOIN employes e ON c.employe_id = e.id
        WHERE c.statut = 'approuvé'
          AND EXTRACT(YEAR FROM c.date_debut) = %s
          AND {scope_where}
        ORDER BY c.date_debut
    """, [annee] + scope_params)
    conges = cur.fetchall()
    cur.close(); conn.close()

    # Total des jours approuvés sur l'année
    total_approuves = sum((c['nombre_jours'] or 0) for c in conges)

    # Répartition par mois (1..12)
    mois_noms = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
    mois_list = [{'nom': mois_noms[m - 1], 'conges': []} for m in range(1, 13)]
    for c in conges:
        debut = c['date_debut']
        # date (objet datetime.date) ou chaîne 'YYYY-MM-DD'
        mois = debut.month if hasattr(debut, 'month') else int(str(debut)[5:7])
        mois_list[mois - 1]['conges'].append(c)

    return render_template('calendrier_conges.html',
                           annee=annee,
                           total_approuves=total_approuves,
                           mois_list=mois_list)

@app.route('/employes/add', methods=['GET','POST'])
@role_required('admin')
def add_employee_alt():
    return redirect(url_for('index'))


# ==================== PARC MATÉRIEL / INVENTAIRES / MAINTENANCE ==============
# Routes et helpers extraits dans blueprints/parc.py (Blueprint ``parc``).


# Blueprints métier : dépendances partagées injectées explicitement pour éviter
# les imports circulaires avec l'application historique.
app.register_blueprint(creer_blueprint_conges({
    'get_db': get_db, 'get_cursor': get_cursor, 'db_cursor': db_cursor,
    'login_required': login_required, 'role_required': role_required,
    'department_scope_sql': department_scope_sql,
    'get_solde_conges': get_solde_conges,
    'get_current_employee': get_current_employee,
    'pagination_info': pagination_info, 'page_list': page_list,
    'peut_decider_rh': _peut_decider_rh,
    'peut_donner_avis': _peut_donner_avis,
    'envoyer_roles': _envoyer_roles,
    'libelle_employe': _libelle_employe,
    'managers_du_departement': _managers_du_departement,
    'notifier_roles': _notifier_roles,
    'user_id_de_employe': _user_id_de_employe,
    'notifier_employe_evenement': _notifier_employe_evenement,
    'create_notification': create_notification, 'queue_email': queue_email,
    'log_action': log_action, 'recalculer_solde': recalculer_solde,
    'DEMANDE_LIBELLES': DEMANDE_LIBELLES,
    'DEMANDE_OUVERTES': DEMANDE_OUVERTES,
    'DEMANDE_EN_ATTENTE': DEMANDE_EN_ATTENTE,
    'DEMANDE_AVIS_RENDU': DEMANDE_AVIS_RENDU,
    'DEMANDE_APPROUVEE': DEMANDE_APPROUVEE,
    'DEMANDE_REFUSEE': DEMANDE_REFUSEE,
    'DEMANDE_ANNULEE': DEMANDE_ANNULEE,
    'Q_SOLDE': Q_SOLDE,
}))

app.register_blueprint(creer_blueprint_recrutement({
    'db_cursor': db_cursor,
    'login_required': login_required,
    'role_required': role_required,
    'get_department_scope': get_department_scope,
    'detect_file_type': detect_file_type,
    'object_storage': object_storage,
    'create_notification': create_notification,
    'queue_email': queue_email,
    'log_action': log_action,
}))

contrats_bp, contrats_api = creer_blueprint_contrats({
    'db_cursor': db_cursor,
    'login_required': login_required,
    'role_required': role_required,
    'get_current_employee': get_current_employee,
    'detect_file_type': detect_file_type,
    'create_notification': create_notification,
    'queue_email': queue_email,
    'log_action': log_action,
    'object_storage': object_storage,
})
job_alertes_contrats = contrats_api['job_alertes_contrats']
app.register_blueprint(contrats_bp)

app.register_blueprint(creer_blueprint_dashboard({
    'db_cursor': db_cursor,
    'login_required': login_required,
    'get_department_scope': get_department_scope,
    'department_scope_sql': department_scope_sql,
    'calculer_retard': calculer_retard,
}))

app.register_blueprint(creer_blueprint_dashboards_roles({
    'db_cursor': db_cursor,
    'login_required': login_required,
    'role_required': role_required,
    'department_scope_sql': department_scope_sql,
    'get_department_scope': get_department_scope,
}))

app.register_blueprint(creer_blueprint_rapports_parc({
    'db_cursor': db_cursor,
    'login_required': login_required,
    'department_scope_sql': department_scope_sql,
    'get_department_scope': get_department_scope,
}))

app.register_blueprint(creer_blueprint_departs({
    'db_cursor': db_cursor,
    'login_required': login_required,
    'role_required': role_required,
    'create_notification': create_notification,
    'queue_email': queue_email,
    'log_action': log_action,
}))

app.register_blueprint(creer_blueprint_auth({
    'db_cursor': db_cursor,
    'login_required': login_required,
    'get_current_user_row': get_current_user_row,
    'get_current_employee': get_current_employee,
    'enregistrer_session': enregistrer_session,
    'cloturer_session': cloturer_session,
    'log_action': log_action,
    'get_role_label': get_role_label,
    'enregistrer_photo_profil': enregistrer_photo_profil,
    'supprimer_photo_profil': supprimer_photo_profil,
    'avatar_folder': AVATAR_FOLDER,
    'limiter': limiter,
}))

app.register_blueprint(creer_blueprint_utilisateurs({
    'db_cursor': db_cursor,
    'login_required': login_required,
    'role_required': role_required,
    'log_action': log_action,
    'get_role_label': get_role_label,
    'create_notification': create_notification,
    'queue_email': queue_email,
    'session_online_window': SESSION_ONLINE_WINDOW_MIN,
    'permanent_session_lifetime': app.config['PERMANENT_SESSION_LIFETIME'],
    'logger': logger,
}))

app.register_blueprint(creer_blueprint_presences({
    'get_db': get_db,
    'get_cursor': get_cursor,
    'db_cursor': db_cursor,
    'login_required': login_required,
    'role_required': role_required,
    'department_scope_sql': department_scope_sql,
    'calculer_retard': calculer_retard,
    'send_retard_email': send_retard_email,
    'notifier_employe_evenement': _notifier_employe_evenement,
    'pagination_info': pagination_info,
    'page_list': page_list,
}))

app.register_blueprint(creer_blueprint_departements({
    'get_db': get_db,
    'get_cursor': get_cursor,
    'login_required': login_required,
    'role_required': role_required,
    'department_scope_sql': department_scope_sql,
}))

app.register_blueprint(creer_blueprint_documents({
    'db_cursor': db_cursor,
    'get_db': get_db,
    'get_cursor': get_cursor,
    'login_required': login_required,
    'role_required': role_required,
    'get_current_employee': get_current_employee,
    'allowed_file': allowed_file,
    'detect_file_type': detect_file_type,
    'log_action': log_action,
    'notifier_employe_evenement': _notifier_employe_evenement,
    'department_scope_sql': department_scope_sql,
    'upload_folder': app.config['UPLOAD_FOLDER'],
    'seuil_expiration': SEUIL_ALERTE_EXPIRATION_DOCUMENTS_JOURS,
    'object_storage': object_storage,
}))

parc_bp, parc_api = creer_blueprint_parc({
    'db_cursor': db_cursor,
    'login_required': login_required,
    'role_required': role_required,
    'create_notification': create_notification,
    'notifier_roles': _notifier_roles,
    'log_action': log_action,
    'get_current_employee': get_current_employee,
    'user_id_de_employe': _user_id_de_employe,
    'pagination_info': pagination_info,
    'page_list': page_list,
    'department_scope_sql': department_scope_sql,
    'get_department_scope': get_department_scope,
    'department_access_denied': _department_access_denied,
    'logger': logger,
})
# Compatibilité des tests et des jobs qui appelaient ce helper depuis app.py.
_notifier_stock_bas = parc_api['notifier_stock_bas']
app.register_blueprint(parc_bp)

recherche_bp, recherche_api = creer_blueprint_recherche({
    'db_cursor': db_cursor,
    'login_required': login_required,
    'department_scope_sql': department_scope_sql,
    'get_role_label': get_role_label,
    'logger': logger,
})
recherche_globale = recherche_api['recherche_globale']
app.register_blueprint(recherche_bp)

app.register_blueprint(creer_blueprint_notifications({
    'login_required': login_required,
    'get_all_notifications': get_all_notifications,
    'mark_all_read': mark_all_read,
}))

app.register_blueprint(creer_blueprint_absences({
    'db_cursor': db_cursor, 'login_required': login_required,
    'role_required': role_required,
    'department_scope_sql': department_scope_sql,
    'pagination_info': pagination_info, 'page_list': page_list,
    'notifier_employe_evenement': _notifier_employe_evenement,
    'generer_absences_automatiques': generer_absences_automatiques,
    'get_department_scope': get_department_scope,
    'ABSENCE_STATUT_LABELS': ABSENCE_STATUT_LABELS,
    'ABSENCE_ACCEPTEE': ABSENCE_ACCEPTEE,
}))

app.register_blueprint(creer_blueprint_justifications({
    'db_cursor': db_cursor,
    'login_required': login_required,
    'role_required': role_required,
    'get_current_employee': get_current_employee,
    'detect_file_type': detect_file_type,
    'create_notification': create_notification,
    'queue_email': queue_email,
    'log_action': log_action,
    'object_storage': object_storage,
}))

app.register_blueprint(creer_blueprint_messagerie({
    'db_cursor': db_cursor,
    'login_required': login_required,
    'detect_file_type': detect_file_type,
    'create_notification': create_notification,
    'queue_email': queue_email,
    'log_action': log_action,
    'department_scope_sql': department_scope_sql,
    'object_storage': object_storage,
}))

# Commandes d'exploitation : bootstrap transitoire du schéma historique,
# migrations Alembic (`flask db ...`) et migration progressive des fichiers.
import click
from services.storage_migration import register_storage_cli


@app.cli.command('bootstrap-db')
def bootstrap_db_command():
    """Amorce idempotemment une base historique avant le premier db upgrade."""
    init_db()
    click.echo('Bootstrap PostgreSQL terminé.')


register_storage_cli(app, db_cursor, object_storage)
health_live, health_ready = register_observability(
    app, get_db, object_storage, alembic_revision='20260814_recrutement'
)
limiter.exempt(health_live)
limiter.exempt(health_ready)

# En production, Render lance `flask bootstrap-db && flask db upgrade` dans le
# preDeployCommand. Aucun worker web ne modifie alors le schéma au démarrage.
if app.config['AUTO_INIT_DB']:
    init_db()
demarrer_scheduler()

if __name__ == '__main__':
    debug_mode = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    if debug_mode and os.environ.get('FLASK_ENV') == 'production':
        raise RuntimeError("FLASK_DEBUG ne doit jamais être activé en production (FLASK_ENV=production).")
    # Développement local uniquement. En production Render : gunicorn app:app
    # --bind 0.0.0.0:$PORT --workers 2 --threads 4
    app.run(debug=debug_mode, host='0.0.0.0', port=5000, threaded=True)